import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
import networkx as nx
from pandas.tseries.offsets import DateOffset
from calendar import monthrange
from openpyxl.utils import get_column_letter
import math
from datetime import datetime, timedelta


"""def identificar_outliers(df):
    Q1 = df['Consumo Total'].quantile(0.25)
    Q3 = df['Consumo Total'].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    # Devuelve una Serie con la etiqueta 'SI' para outliers y 'NO' para no outliers
    return pd.Series(np.where((df['Consumo Total'] < lower_bound) | (df['Consumo Total'] > upper_bound), 'SI', 'NO'), index=df.index)
"""
def identificar_outliers(x):
    Q1 = x.quantile(0.25)
    Q3 = x.quantile(0.75)
    IQR = Q3 - Q1
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR  
    return np.where((x < lower) | (x > upper), 'SI', 'NO')


def quitar_outliers_v2(grupo):

    # CÁLCULOS BASE
    media = grupo.mean()
    desv_est = grupo.std()
   
    # COEFICIENTE DE VARIACIÓN
    cv = desv_est / media

    # MÉTODO TOLERANCIA
    if cv < 0.7:

        margen = media * 0.70
        li = media - margen
        ls = media + margen
        return grupo.apply(
            lambda x:
            "NO"
            if li <= x <= ls
            else "SI"
        )

    # MÉTODO IQR
    else:

        q1 = grupo.quantile(0.25)
        q3 = grupo.quantile(0.75)
        iqr = q3 - q1
        li = q1 - 1.5 * iqr
        ls = q3 + 1.5 * iqr
        return grupo.apply(
            lambda x:
            "NO"
            if li <= x <= ls
            else "SI"
        )

def limpieza_consumos(ResumenPrevKardexSAPcliente,ResumenPrevKardexwhtAlmacen,df_componente_consumo,reference_period,Linea_Negocio):
    start_date = reference_period - pd.DateOffset(months=12)##12
    #ResumenPrevKardexwhtAlmacen=df_componente_consumo.copy()
    if Linea_Negocio == "MANGUERA Y CONEXION":
        ResumenPrevKardexSAPcliente['month_year'] = pd.to_datetime( ResumenPrevKardexSAPcliente['month_year'])       
        df_filtrado = ResumenPrevKardexSAPcliente[
            (ResumenPrevKardexSAPcliente['month_year'] >= start_date) &
            (ResumenPrevKardexSAPcliente['month_year'] <= reference_period)
        ]
        df_sin_apm = df_filtrado[df_filtrado['Cliente_Final'] != "APM TERMINALS CALLAO SOCIEDAD ANONIMA"]
        df_componente_consumo = df_sin_apm.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
        df_ceros = df_componente_consumo[ df_componente_consumo['Consumo Total'] == 0].copy()    
        df_no_ceros = df_componente_consumo[df_componente_consumo['Consumo Total'] > 0].copy()     
        df_no_ceros["Outlier"] = (df_no_ceros.groupby("SAP")["Consumo Total"].transform(quitar_outliers_v2))
           
        #Sacamos consumos considerados
        df_consumos_f = pd.concat([df_no_ceros, df_ceros], ignore_index=True)
        df_consumos_f['Outlier'] = df_consumos_f['Outlier'].fillna('NO')
        df_consumos_f = df_consumos_f.sort_values('month_year').reset_index(drop=True)
        """df_consumos_f = df_consumos_f.rename(columns={
            'SAP': 'Componente',
            'Consumo Total': 'Consumo Componente'
        })"""
        
    else:
        ResumenPrevKardexwhtAlmacen['month_year'] = pd.to_datetime( ResumenPrevKardexwhtAlmacen['month_year'])       
        df_filtrado = ResumenPrevKardexwhtAlmacen[
            (ResumenPrevKardexwhtAlmacen['month_year'] >= start_date) &
            (ResumenPrevKardexwhtAlmacen['month_year'] <= reference_period)
        ]
        ResumenPrevKardexwhtAlmacen = df_filtrado.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
        df_ceros = ResumenPrevKardexwhtAlmacen[ResumenPrevKardexwhtAlmacen['Consumo Total'] == 0 ].copy()
    
        df_no_ceros = ResumenPrevKardexwhtAlmacen[
            ResumenPrevKardexwhtAlmacen['Consumo Total'] > 0
        ].copy() 
        df_no_ceros["Outlier"] = ( df_no_ceros.groupby("SAP")["Consumo Total"].transform(quitar_outliers_v2)        )       
        #Sacamos consumos considerados
        df_consumos_f = pd.concat([df_no_ceros, df_ceros], ignore_index=True)
        df_consumos_f['Outlier'] = df_consumos_f['Outlier'].fillna('NO')
        df_consumos_f = df_consumos_f.sort_values('month_year').reset_index(drop=True)
                             
    return df_consumos_f   

def consumo_24M(ResumenPrevKardexSAPcliente,ResumenPrevKardexwhtAlmacen,reference_period, Linea_Negocio):
  
    if Linea_Negocio == "MANGUERA Y CONEXION":

        df_history = ResumenPrevKardexSAPcliente.copy()
        df_history['month_year'] = pd.to_datetime(df_history['month_year'])    
        df_history = df_history[ df_history['Cliente_Final'] != "APM TERMINALS CALLAO SOCIEDAD ANONIMA" ]
    else:
        df_history = ResumenPrevKardexwhtAlmacen.copy()
        df_history['month_year'] = pd.to_datetime(df_history['month_year'])
    
    df_history = (df_history .groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index())
    df_ceros = df_history[ df_history['Consumo Total'] == 0].copy()
    df_no_ceros = df_history[df_history['Consumo Total'] > 0].copy()
    df_no_ceros["Outlier"] = (df_no_ceros.groupby("SAP")["Consumo Total"].transform(quitar_outliers_v2))

    df_history = pd.concat( [df_no_ceros, df_ceros],  ignore_index=True)
    df_history['Outlier'] = (df_history['Outlier'] .fillna('NO'))
    df_history = (df_history.sort_values(by=['SAP', 'month_year'], ascending=[True, True]) .reset_index(drop=True))
    #df_history = (df_history .sort_values(['SAP', 'month_year']) .reset_index(drop=True))
    return df_history

def limpieza_consumos_setting(ResumenPrevKardexwhtAlmacen,df_componente_consumo,dfMRP_filter,reference_period):
    ##
    df_history = df_componente_consumo.copy()
    df_history['month_year'] = pd.to_datetime(df_history['month_year'])
    df_history = (df_history.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index())
    df_ceros_hist = df_history[df_history['Consumo Total'] == 0].copy()
    df_no_ceros_hist = df_history[df_history['Consumo Total'] > 0].copy()
    df_no_ceros_hist["Outlier"] = (df_no_ceros_hist.groupby("SAP")["Consumo Total"].transform(quitar_outliers_v2))
    df_history = pd.concat([df_no_ceros_hist, df_ceros_hist],ignore_index=True)
    df_history['Outlier'] = (df_history['Outlier'].fillna('NO'))
    df_history = (df_history.sort_values(['SAP', 'month_year']).reset_index(drop=True))
    
    df_hist_pf = ResumenPrevKardexwhtAlmacen.copy()
    df_hist_pf['month_year'] = pd.to_datetime(df_hist_pf['month_year'])
    df_hist_pf = (df_hist_pf.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index())
    df_ceros_pf = df_hist_pf[ df_hist_pf['Consumo Total'] == 0].copy()
    df_no_ceros_pf = df_hist_pf[ df_hist_pf['Consumo Total'] > 0].copy() 
    df_no_ceros_pf["Outlier"] = ( df_no_ceros_pf.groupby("SAP")["Consumo Total"] .transform(quitar_outliers_v2))       
    df_hist_pf = pd.concat([df_no_ceros_pf, df_ceros_pf], ignore_index=True)
    df_hist_pf['Outlier'] = df_hist_pf['Outlier'].fillna('NO')
    df_hist_pf = df_hist_pf.sort_values(['SAP','month_year']).reset_index(drop=True)  
    df_hist_pf = df_hist_pf[ df_hist_pf["SAP"].isin(dfMRP_filter.index)]
    df_history = pd.concat( [df_hist_pf, df_history], ignore_index=True)
    df_history = ( df_history.sort_values(['SAP', 'month_year']).reset_index(drop=True))
    ##
    start_date = reference_period - pd.DateOffset(months=12)
    start_date2 = reference_period - pd.DateOffset(months=24)
    #Consumos componentes
    df_componente_consumo['month_year'] = pd.to_datetime( df_componente_consumo['month_year'])       
    df_filtrado = df_componente_consumo[
        (df_componente_consumo['month_year'] >= start_date) &
        (df_componente_consumo['month_year'] <= reference_period)
    ]
    df_componente_consumo = df_filtrado.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
    df_ceros = df_componente_consumo[
        df_componente_consumo['Consumo Total'] == 0
    ].copy()

    df_no_ceros = df_componente_consumo[
        df_componente_consumo['Consumo Total'] > 0
    ].copy() 
    df_no_ceros["Outlier"] = (
        df_no_ceros.groupby("SAP")["Consumo Total"]
        .transform(quitar_outliers_v2)
    )       
    #Sacamos consumos considerados
    df_consumos_f = pd.concat([df_no_ceros, df_ceros], ignore_index=True)
    df_consumos_f['Outlier'] = df_consumos_f['Outlier'].fillna('NO')
    df_consumos_componentes = df_consumos_f.sort_values('month_year').reset_index(drop=True)
    
    #Consumos producto finales
    ResumenPrevKardexwhtAlmacen['month_year'] = pd.to_datetime( ResumenPrevKardexwhtAlmacen['month_year'])       
    df_filtrado = ResumenPrevKardexwhtAlmacen[
        (ResumenPrevKardexwhtAlmacen['month_year'] >= start_date) &
        (ResumenPrevKardexwhtAlmacen['month_year'] <= reference_period)
    ]
    df_filtrado = df_filtrado.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
    df_ceros = df_filtrado[
        df_filtrado['Consumo Total'] == 0
    ].copy()

    df_no_ceros = df_filtrado[
        df_filtrado['Consumo Total'] > 0
    ].copy() 
    df_no_ceros["Outlier"] = (
        df_no_ceros.groupby("SAP")["Consumo Total"]
        .transform(quitar_outliers_v2)
    )       
    #Sacamos consumos considerados
    df_consumos = pd.concat([df_no_ceros, df_ceros], ignore_index=True)
    df_consumos['Outlier'] = df_consumos['Outlier'].fillna('NO')
    df_consumos = df_consumos.sort_values('month_year').reset_index(drop=True)
    df_consumos_standar = df_consumos.copy()
    df_consumos_producto_final = df_consumos[df_consumos["SAP"].isin(dfMRP_filter.index)]    
    df_consumos_pf_comp_12M = pd.concat([df_consumos_producto_final, df_consumos_componentes], ignore_index=True)
    
    #######################
    #Historia didactica 24M
    #Consumos componentes
    df_componente_consumo['month_year'] = pd.to_datetime( df_componente_consumo['month_year'])       
    df_filtrado = df_componente_consumo[
        (df_componente_consumo['month_year'] >= start_date2) &
        (df_componente_consumo['month_year'] <= reference_period)
    ]
    df_componente_consumo2 = df_filtrado.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
    #Consumos producto finales
    ResumenPrevKardexwhtAlmacen['month_year'] = pd.to_datetime( ResumenPrevKardexwhtAlmacen['month_year'])       
    df_filtrado = ResumenPrevKardexwhtAlmacen[
        (ResumenPrevKardexwhtAlmacen['month_year'] >= start_date2) &
        (ResumenPrevKardexwhtAlmacen['month_year'] <= reference_period)
    ]
    df_filtrado = df_filtrado.groupby(['SAP', 'month_year'])['Consumo Total'].sum().reset_index()
    df_consumos_producto_final2 = df_filtrado[df_filtrado["SAP"].isin(dfMRP_filter.index)] 
    df_consumos_pf_comp_24M = pd.concat([df_consumos_producto_final2, df_componente_consumo2], ignore_index=True)
    df_consumos_pf_comp_24M = df_consumos_pf_comp_24M.merge(
    df_consumos_pf_comp_12M[["SAP", "month_year", "Outlier"]],
    on=["SAP", "month_year"],
    how="left"
    )
    df_consumos_pf_comp_24M['Outlier'] = df_consumos_pf_comp_24M['Outlier'].fillna('NO')
                             
    return df_consumos_componentes , df_history,  df_consumos_standar #df_consumos_pf_comp_24M ,

def agregar_consumo_promedio(dfActiveCode, ResumenPrevKardexSAPcliente,reference_period):
    
    start_date = reference_period - pd.DateOffset(months=12)
    ResumenPrevKardexSAPcliente = ResumenPrevKardexSAPcliente[
        (ResumenPrevKardexSAPcliente['month_year'] >= start_date) &
        (ResumenPrevKardexSAPcliente['month_year'] <= reference_period)
    ]    
    consumos = (ResumenPrevKardexSAPcliente.groupby('SAP')['Consumo Total'] .mean().reset_index()  )
    consumos.rename(columns={'SAP': 'Codigo_SAP','Consumo Total': 'Consumo Promedio 12M'}, inplace=True)
    dfActiveCode = dfActiveCode.merge(consumos, on='Codigo_SAP', how='left')
    dfActiveCode['Consumo Promedio 12M'] = dfActiveCode['Consumo Promedio 12M'].fillna(0).round(1) 
    
    return dfActiveCode   

def agregar_consumo_promedio_cliente(dfActiveCode, ResumenPrevKardexSAPcliente, reference_period):

    # Asegurar formato datetime
    ResumenPrevKardexSAPcliente['month_year'] = pd.to_datetime(
        ResumenPrevKardexSAPcliente['month_year']
    )

    # Rango últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    df_filtrado = ResumenPrevKardexSAPcliente[
        (ResumenPrevKardexSAPcliente['month_year'] >= start_date) &
        (ResumenPrevKardexSAPcliente['month_year'] <= reference_period)
    ].copy()

    # Agrupar por SAP y Cliente
    df_grouped = (
        df_filtrado
        .groupby(['SAP', 'Cliente_Final'])
        .agg({
            'Consumo Total': 'mean',
            'Valor Total Soles': 'sum'
        })
        .reset_index()
    )

    # Redondear promedio a 1 decimal
    df_grouped['Consumo Total'] = df_grouped['Consumo Total'].round(1)
    # Excluir clientes con promedio 0
    df_grouped = df_grouped[df_grouped['Consumo Total'] > 0]

    # Ordenar por Valor Total Soles (desc)
    df_grouped = df_grouped.sort_values(
        by=['SAP', 'Valor Total Soles'],
        ascending=[True, False]
    )

    # Construir texto final por SAP
    def construir_texto(grupo):
        return " // ".join(
            grupo['Cliente_Final'] + " (" + grupo['Consumo Total'].astype(str) + ")"
        )

    df_texto = (
        df_grouped
        .groupby('SAP')
        .apply(construir_texto)
        .reset_index()
        .rename(columns={
            'SAP': 'Codigo_SAP',
            0: 'Consumo Prom 12M Cliente'
        })
    )

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(
        df_texto,
        on='Codigo_SAP',
        how='left'
    )

    # Si no tuvo consumo en 12M → cadena vacía
    dfActiveCode['Consumo Prom 12M Cliente'] = (
        dfActiveCode['Consumo Prom 12M Cliente'].fillna('')
    )

    return dfActiveCode

def agregar_consumo_promedio_clienteMNG(dfActiveCode, ResumenPrevKardexSAPcliente, reference_period):

    # Asegurar formato datetime
    ResumenPrevKardexSAPcliente['month_year'] = pd.to_datetime(
        ResumenPrevKardexSAPcliente['month_year']
    )

    # Rango últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    df_filtrado = ResumenPrevKardexSAPcliente[
        (ResumenPrevKardexSAPcliente['month_year'] >= start_date) &
        (ResumenPrevKardexSAPcliente['month_year'] <= reference_period)
    ].copy()

    # Agrupar por SAP y Cliente
    df_grouped = (
        df_filtrado
        .groupby(['SAP', 'Cliente_Final'])
        .agg({
            'Consumo Total': 'mean',
            'Valor Total Soles': 'sum'
        })
        .reset_index()
    )

    # Redondear promedio
    df_grouped['Consumo Total'] = df_grouped['Consumo Total'].round(1)
    df_grouped = df_grouped[df_grouped['Consumo Total'] > 0]

    # Lista 
    clientes_objetivo = [
       "APM TERMINALS CALLAO SOCIEDAD ANONIMA",
        "MINERA LAS BAMBAS S.A.",
        "KOMATSU-MITSUI MAQUINARIAS PERU S.A.",
        "MINERA CHINALCO PERÚ S.A.",
        "COSAPI MINERIA S.A.C."
    ]
    
    df_grouped['Cliente_Agrupado'] = df_grouped['Cliente_Final'].apply(lambda x: x if x in clientes_objetivo else 'OTROS')

    df_grouped_final = (
        df_grouped
        .groupby(['SAP', 'Cliente_Agrupado'])
        .agg({'Consumo Total': 'mean'})
        .reset_index()
    )

    # Pivotear
    df_pivot = df_grouped_final.pivot_table(
        index='SAP',
        columns='Cliente_Agrupado',
        values='Consumo Total',
        aggfunc='first'
    ).reset_index()
    df_pivot.columns.name = None
    # Renombrar SAP para merge
    df_pivot = df_pivot.rename(columns={'SAP': 'Codigo_SAP'})

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(
        df_pivot,
        on='Codigo_SAP',
        how='left'
    )

    columnas = df_pivot.columns.drop('Codigo_SAP')
    dfActiveCode[columnas] = (dfActiveCode[columnas].apply(pd.to_numeric, errors='coerce').fillna(0).round(1))
    dfActiveCode = dfActiveCode.drop(columns=["APM TERMINALS CALLAO SOCIEDAD ANONIMA"], errors='ignore')

    return dfActiveCode



def agregar_consumo_prom_compo(dfActiveCode, df_componente_consumo, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    df_componente_consumo['month_year'] = pd.to_datetime(df_componente_consumo['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = df_componente_consumo[
        (df_componente_consumo['month_year'] >= start_date) &
        (df_componente_consumo['month_year'] <= reference_period)
    ]

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'Consumo Promedio Comp 12M'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['Consumo Promedio Comp 12M'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_promedio_forecast_3m_mineria(dfActiveCode, dfForecastProm3M, dfBOMfinal):
    # Merge directo de Promedio_3m
    dfActiveCode = dfActiveCode.merge(
        dfForecastProm3M, on="Codigo_SAP", how="left"
    )

    # Si no se encuentra Promedio_3m, usar Consumo Promedio
    dfActiveCode["Promedio_3m_Forecast"] = dfActiveCode["Promedio_3m_Forecast"].fillna(
        dfActiveCode["Consumo Promedio 12M"]
    )

    # === Calcular Promedio_3m-C (equivalencia de componentes) ===
    # Unir dfForecastProm3M con BOM
    df_promC = dfForecastProm3M.merge(dfBOMfinal, on="Codigo_SAP", how="inner")

    # Calcular equivalencia
    df_promC["Promedio_3m_Forecast-C"] = df_promC["Promedio_3m_Forecast"] * df_promC["Q"]

    # Agrupar por Componente
    df_promC = df_promC.groupby("Componente")["Promedio_3m_Forecast-C"].sum().reset_index()

    # Renombrar Componente como Codigo_SAP para el merge
    df_promC.rename(columns={"Componente": "Codigo_SAP"}, inplace=True)

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(df_promC, on="Codigo_SAP", how="left")

    # Si no se encuentra Promedio_3m-C, usar Consumo Promedio Comp
    dfActiveCode["Promedio_3m_Forecast-C"] = dfActiveCode["Promedio_3m_Forecast-C"].fillna(
        dfActiveCode["Consumo Promedio Comp 12M"]
    )

    return dfActiveCode



def agregar_fixedcost(dfActiveCode, dfCostosImportacion):


    # Filtrar solo los datos dentro del rango de 12 meses
    dffixedcostfilt = dfCostosImportacion[
        (dfCostosImportacion['Modalidad'] == "MARITIMO") &
        (dfCostosImportacion['Outlier'] == "NO")
    ]
    

    # Calcular el consumo promedio por Código SAP
    dffixedcostfilt = dffixedcostfilt.groupby('Número de artículo')['Costo Fijo Total'].mean().reset_index()
    dffixedcostfilt.rename(columns={'Número de artículo': 'Codigo_SAP', 'Costo Fijo Total': 'Fixed Cost Prom'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(dffixedcostfilt, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['Fixed Cost Prom'].fillna(0, inplace=True)

    return dfActiveCode

def agregar_valor_consumo(dfActiveCode, ResumenPrevKardexSAPsinAlmacen, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAPsinAlmacen['month_year'] = pd.to_datetime(ResumenPrevKardexSAPsinAlmacen['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAPsinAlmacen[
        (ResumenPrevKardexSAPsinAlmacen['month_year'] >= start_date) &
        (ResumenPrevKardexSAPsinAlmacen['month_year'] <= reference_period)
    ]

    # Calcular la suma de "Valor Total Soles" por Código SAP
    valor_consumo = df_filtrado.groupby('SAP')['Valor Total Soles'].sum().reset_index()
    valor_consumo.rename(columns={'SAP': 'Codigo_SAP', 'Valor Total Soles': 'Valor de Consumo'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(valor_consumo, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo valor de consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['Valor de Consumo'].fillna(0, inplace=True)

    return dfActiveCode

def agregar_valor_consumoHYD(dfActiveCode, ResumenPrevKardexSAPsinAlmacen, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAPsinAlmacen['month_year'] = pd.to_datetime(ResumenPrevKardexSAPsinAlmacen['month_year'])

    # Calcular la suma de "Valor Total Soles" por Código SAP
    valor_consumo = ResumenPrevKardexSAPsinAlmacen.groupby('SAP')['Valor Total Soles'].sum().reset_index()
    valor_consumo.rename(columns={'SAP': 'Codigo_SAP', 'Valor Total Soles': 'Valor de Consumo'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(valor_consumo, on='Codigo_SAP', how='left')

     # Reemplazar NaN con 0 y valores negativos con 0
    dfActiveCode['Valor de Consumo'] = dfActiveCode['Valor de Consumo'].fillna(0).clip(lower=0)
    
    return dfActiveCode


def agregar_consumo_promedio200(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '200']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgKMMP'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgKMMP'].fillna(0, inplace=True)

    return dfActiveCode

def agregar_consumo_promedio330(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '330']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgCOS'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgCOS'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedio300(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '300']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgBAMB'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgBAMB'].fillna(0, inplace=True)

    return dfActiveCode

def agregar_consumo_promedio233(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '233']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgQUELL'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgQUELL'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedio310(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '310']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgANTA'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgANTA'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedio231(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '231']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgSOUT'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgSOUT'].fillna(0, inplace=True)

    return dfActiveCode

def agregar_consumo_promedio360(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '360']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgCHI'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgCHI'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedio385(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])
    start_date = reference_period - pd.DateOffset(months=11)
    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    # Filtrar solo almacén 385
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == '385']

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={ 'SAP': 'Codigo_SAP',  'Consumo Total': 'AvgTIAMARIA' }, inplace=True)
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')
    dfActiveCode['AvgTIAMARIA'].fillna(0, inplace=True)

    return dfActiveCode



def agregar_consumo_promedioAPM(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    # Convertir month_year a tipo datetime para facilitar el filtrado
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'] == 'APM-01']
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgAPM-01'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgAPM-01'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedioSucursales(dfActiveCode, ResumenPrevKardexSAP, reference_period):

    # Convertir month_year a datetime
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir sucursales
    sucursales = {'210', '220', '230', '232', '240', '241'}

    # Definir rango de últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar últimos 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]

    # Filtrar solo sucursales
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'].isin(sucursales)]

    # Calcular consumo promedio mensual por SAP
    consumo_promedio = (df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index())

    # Renombrar columnas para el merge
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP','Consumo Total': 'AvgSucursales'},inplace=True)
    dfActiveCode = dfActiveCode.merge(consumo_promedio,on='Codigo_SAP', how='left')
    # Rellenar NaN con 0
    dfActiveCode['AvgSucursales'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_consumo_promedioCALLAO(dfActiveCode, ResumenPrevKardexSAP, reference_period):
    ResumenPrevKardexSAP['Código de almacén'] = ResumenPrevKardexSAP['Código de almacén'].astype(str)
    # Convertir month_year a tipo datetime para facilitar el filtrado
    almacenes_calloycd = ['100', '101', '102', '105', '106', '110', '114', '400', '495', '395', '500', '502', '503']
    ResumenPrevKardexSAP['month_year'] = pd.to_datetime(ResumenPrevKardexSAP['month_year'])

    # Definir el rango de fechas para los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar solo los datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAP[
        (ResumenPrevKardexSAP['month_year'] >= start_date) &
        (ResumenPrevKardexSAP['month_year'] <= reference_period)
    ]
    
    df_filtrado = df_filtrado[df_filtrado['Código de almacén'].isin(almacenes_calloycd)]
    

    # Calcular el consumo promedio por Código SAP
    consumo_promedio = df_filtrado.groupby('SAP')['Consumo Total'].mean().reset_index()
    consumo_promedio.rename(columns={'SAP': 'Codigo_SAP', 'Consumo Total': 'AvgCALLAO'}, inplace=True)

    # Hacer un merge eficiente con dfActiveCode para agregar la nueva columna
    dfActiveCode = dfActiveCode.merge(consumo_promedio, on='Codigo_SAP', how='left')

    # Si algún código SAP no tuvo consumo en los últimos 12 meses, llenamos con 0 en lugar de NaN
    dfActiveCode['AvgCALLAO'].fillna(0, inplace=True)

    return dfActiveCode

###########################################
def generar_transito_por_fechas(dfActiveCode, FreservSAP_filtrado, SeguiBOSAP_f):

    FreservSAP_filtrado = FreservSAP_filtrado[FreservSAP_filtrado["Sociedad"] == "MP"]
    SeguiBOSAP_f = SeguiBOSAP_f[SeguiBOSAP_f["Sociedad"] == "MP"]

    FreservSAP_filtrado["Restante"] = (
    FreservSAP_filtrado["Restante"]
    .astype(str)  # Convertir todos los valores a string
    .str.replace(',', '', regex=False)  # Remover separadores de miles (coma)
    .replace(r'^\s*$', None, regex=True)  # Reemplazar valores vacíos o espacios con None (para que sean NaN)
    .astype(float)  # Convertir finalmente a float
    )
    
    SeguiBOSAP_f["Pendiente"] = (
    SeguiBOSAP_f["Pendiente"]
    .astype(str)  # Convertir todos los valores a string
    .str.replace(',', '', regex=False)  # Remover separadores de miles (coma)
    .replace(r'^\s*$', None, regex=True)  # Reemplazar valores vacíos o espacios con None (para que sean NaN)
    .astype(float)  # Convertir finalmente a float
    )
    
    FreservSAP_filtrado["ETA Marco"] = pd.to_datetime(FreservSAP_filtrado["ETA Marco"], errors="coerce")
    SeguiBOSAP_f["ETA Marco"] = pd.to_datetime(SeguiBOSAP_f["ETA Marco"], errors="coerce")
    SeguiBOSAP_f["Fecha Orden"] = pd.to_datetime(SeguiBOSAP_f["Fecha Orden"], errors="coerce")
    SeguiBOSAP_f["Ultima Fecha"] = pd.to_datetime(SeguiBOSAP_f["Ultima Fecha"], errors="coerce")
    
    hoy = pd.Timestamp.today().normalize()
    etas_freserv = FreservSAP_filtrado[["Número de artículo", "ETA Marco"]]
    etas_bo = SeguiBOSAP_f[["Número de artículo", "ETA Marco"]]
    etas_total = pd.concat([etas_freserv, etas_bo], ignore_index=True)
    etas_total = etas_total[(etas_total["ETA Marco"].notna()) &(etas_total["ETA Marco"] >= hoy)]  
    proxima_llegada = (etas_total.groupby("Número de artículo")["ETA Marco"].min())
      
    # Procesar FreservSAP_filtrado (prefijo "T") con formato de fecha corto y redondeo de cantidades
    transito_dict_T = (
        FreservSAP_filtrado
        .groupby("Número de artículo")
        .apply(lambda x: " / ".join(
            f"T({row['PO']},{round(row['Restante'], 1)},{row['ETA Marco'].strftime('%Y-%m-%d') if pd.notna(row['ETA Marco']) else '1900-01-01'},{row['Cat ETA']})"
            for _, row in x.iterrows()
        ))
        .to_dict()
    )

    # Procesar SeguiBOSAP_f (prefijo "BO"), manejando valores nulos en "Ultima Fecha"
    transito_dict_BO = (
    SeguiBOSAP_f
    .groupby("Número de artículo")
    .apply(lambda x: " / ".join(
        f"BO({row['Orden de Compra']},"
        f"{row['Fecha Orden'].strftime('%Y-%m-%d') if pd.notna(row['Fecha Orden']) else 'ND'},"
        f"{round(row['Pendiente'], 1)},"
        f"{row['Ultima Fecha'].strftime('%Y-%m-%d') if pd.notna(row['Ultima Fecha']) else 'ND'},"
        f"{row['ETA Marco'].strftime('%Y-%m-%d') if pd.notna(row['ETA Marco']) else 'ND'},"
        f"{row['Pais de Origen'] if pd.notna(row['Pais de Origen']) else 'ND'})"
        for _, row in x.iterrows()
        ))
        .to_dict()
    )

    # Unir los resultados de ambos diccionarios
    def combinar_transitos(codigo_sap):
        t_part = transito_dict_T.get(codigo_sap, "")
        bo_part = transito_dict_BO.get(codigo_sap, "")
        return " / ".join(filter(None, [t_part, bo_part]))  # Evita dobles separadores si uno está vacío

    # Aplicar la combinación a dfActiveCode
    dfActiveCode["Transito por Fechas"] = dfActiveCode["Codigo_SAP"].map(combinar_transitos)
    # Agregar columna PROXIMA LLEGADA2
    dfActiveCode["Fecha_LLegada"] = (dfActiveCode["Codigo_SAP"].map(proxima_llegada))
    
    return dfActiveCode

# Función para generar la columna "Transito por Fechas" con FreservSAP_filtrado y SeguiBOSAP_f
def calcular_totaltransito(dfActiveCode, FreservSAP_filtrado, SeguiBOSAP_f):
    FreservSAP_filtrado = FreservSAP_filtrado[FreservSAP_filtrado["Sociedad"] == "MP"]
    SeguiBOSAP_f = SeguiBOSAP_f[SeguiBOSAP_f["Sociedad"] == "MP"]
    FreservSAP_filtrado["Restante"] = (
    FreservSAP_filtrado["Restante"]
    .astype(str)  # Convertir todos los valores a string
    .str.replace(',', '', regex=False)  # Remover separadores de miles (coma)
    .replace(r'^\s*$', None, regex=True)  # Reemplazar valores vacíos o espacios con None (para que sean NaN)
    .astype(float)  # Convertir finalmente a float
    )
    
    SeguiBOSAP_f["Pendiente"] = (
    SeguiBOSAP_f["Pendiente"]
    .astype(str)  # Convertir todos los valores a string
    .str.replace(',', '', regex=False)  # Remover separadores de miles (coma)
    .replace(r'^\s*$', None, regex=True)  # Reemplazar valores vacíos o espacios con None (para que sean NaN)
    .astype(float)  # Convertir finalmente a float
    )

    # Agrupar por Codigo_SAP y sumar los valores
    restante_por_codigo = FreservSAP_filtrado.groupby("Número de artículo")["Restante"].sum(min_count=1)
    pendiente_por_codigo = SeguiBOSAP_f.groupby("Número de artículo")["Pendiente"].sum(min_count=1)

    # Combinar ambos en un único DataFrame
    transito_total = pd.DataFrame({
        "Restante": restante_por_codigo,
        "Pendiente": pendiente_por_codigo
    }).fillna(0)
    
    # Renombrar el índice a 'Codigo_SAP'
    transito_total.index.name = "Codigo_SAP"

    # Calcular Transito Total
    transito_total["Transito Total"] = transito_total["Restante"] + transito_total["Pendiente"]

    # Unir con dfActiveCode según Codigo_SAP
    dfActiveCode = dfActiveCode.merge(
        transito_total["Transito Total"], on="Codigo_SAP", how="left"
    ).fillna({"Transito Total": 0})

    return dfActiveCode

def calcular_totaltransito_comp(dfActiveCode, FreservSAP_filtrado, SeguiBOSAP_f, dfBOMfinal):
    
    # Limpieza y conversión de columnas numéricas
    FreservSAP_filtrado["Restante"] = (
        FreservSAP_filtrado["Restante"]
        .astype(str)
        .str.replace(',', '', regex=False)
        .replace(r'^\s*$', None, regex=True)
        .astype(float)
    )
    
    SeguiBOSAP_f["Pendiente"] = (
        SeguiBOSAP_f["Pendiente"]
        .astype(str)
        .str.replace(',', '', regex=False)
        .replace(r'^\s*$', None, regex=True)
        .astype(float)
    )

    # Agrupar por Código SAP y sumar
    restante_por_codigo = FreservSAP_filtrado.groupby("Número de artículo")["Restante"].sum(min_count=1)
    pendiente_por_codigo = SeguiBOSAP_f.groupby("Número de artículo")["Pendiente"].sum(min_count=1)

    # Combinar ambos
    transito_total = pd.DataFrame({
        "Restante": restante_por_codigo,
        "Pendiente": pendiente_por_codigo
    }).fillna(0)

    transito_total.index.name = "Codigo_SAP"
    transito_total["Transito Total"] = transito_total["Restante"] + transito_total["Pendiente"]

    # === Cálculo del Transito Total - C (componentes equivalentes) ===
    # Unir BOM con transito total
    transito_equivalente = transito_total[["Transito Total"]].reset_index().merge(
        dfBOMfinal, left_on="Codigo_SAP", right_on="Codigo_SAP", how="inner"
    )

    # Calcular el tránsito equivalente
    transito_equivalente["Transito Total - C"] = transito_equivalente["Transito Total"] * transito_equivalente["Q"]

    # Agrupar por Componente para acumular el tránsito total de cada componente
    transito_por_componente = transito_equivalente.groupby("Componente")["Transito Total - C"].sum().reset_index()

    # Renombrar columna para hacer merge con dfActiveCode
    transito_por_componente.rename(columns={"Componente": "Codigo_SAP"}, inplace=True)

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(
        transito_total[["Transito Total"]], on="Codigo_SAP", how="left"
    )
    dfActiveCode = dfActiveCode.merge(
        transito_por_componente, on="Codigo_SAP", how="left"
    )

    # Rellenar NaNs con 0
    dfActiveCode[["Transito Total", "Transito Total - C"]] = dfActiveCode[
        ["Transito Total", "Transito Total - C"]
    ].fillna(0)

    return dfActiveCode


def calcular_comprometido_comp(dfActiveCode, dfBOMfinal):
    #dfActiveCode=df_resultado_ltvar.copy()
    # Limpieza y conversión de columnas numéricas
    dfActiveCode["Comprometido"] = (
        dfActiveCode["Comprometido"]
        .astype(str)
        .str.replace(',', '', regex=False)
        .replace(r'^\s*$', None, regex=True)
        .astype(float)
    )
    
    #Ajustamos compromentidos para componentes con setting
    mask = dfActiveCode["Codigo_SAP"].isin(dfBOMfinal["Componente"])
    dfActiveCode.loc[mask, "Comprometido"] = (
    dfActiveCode.loc[mask, "Comprometido"] - 
    dfActiveCode.loc[mask, "En OF"]
    )
    dfActiveCode.loc[mask, "En OF"] =0
    
    dfComprometido = dfActiveCode[['Codigo_SAP', 'Comprometido',"En OF"]]

    comprometido_equivalente = dfComprometido.merge(
        dfBOMfinal, left_on="Codigo_SAP", right_on="Codigo_SAP", how="inner"
    )

    # Calcular el comprometido equivalente
    comprometido_equivalente["Comprometido - C"] = (
        comprometido_equivalente["Comprometido"] * comprometido_equivalente["Q"]
    )

    # Agrupar por Componente para acumular el tránsito total de cada componente
    comprometido_por_componente = (
        comprometido_equivalente.groupby("Componente")["Comprometido - C"]
        .sum()
        .reset_index()
    )

    # Renombrar columna para hacer merge con dfActiveCode
    comprometido_por_componente.rename(columns={"Componente": "Codigo_SAP"}, inplace=True)

    dfActiveCode = dfActiveCode.merge(
        comprometido_por_componente, on="Codigo_SAP", how="left"
    )

    # Rellenar NaNs con 0
    dfActiveCode["Comprometido - C"] = dfActiveCode["Comprometido - C"].fillna(0)

    return dfActiveCode

def agregar_stock_disponible(dfActiveCode, dfConsInvPortalUniconSAP):
    """
    Agrega las columnas 'Stock Disponible' y 'Stock no Disponible' a dfActiveCode.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        dfConsInvPortalUniconSAP (pd.DataFrame): DataFrame con la información de stock y almacenes.

    Returns:
        pd.DataFrame: dfActiveCode con las nuevas columnas agregadas.
    """
    # Definir los almacenes considerados como 'no disponible'
    almacenes_no_disponibles = {'104', '111', '112', '501'}

    
    # Agrupar por SAP y calcular los valores
    stock_por_sap = dfConsInvPortalUniconSAP.groupby("SAP")["Stock"].agg([
        lambda x: x[dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum(),
        lambda x: x[~dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum()
    ]).reset_index()

    # Renombrar columnas
    stock_por_sap.columns = ["Codigo_SAP", "Stock no Disponible", "Stock Disponible"]

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(stock_por_sap, on="Codigo_SAP", how="left").fillna(0)

    return dfActiveCode


def agregar_stock_disponible_comp(dfActiveCode, dfConsInvPortalUniconSAP, dfBOMfinal):
    """
    Agrega las columnas de stock real y stock equivalente por componente.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        dfConsInvPortalUniconSAP (pd.DataFrame): DataFrame con columnas ['SAP', 'Almacen', 'Stock'].
        dfBOMfinal (pd.DataFrame): DataFrame con columnas ['Codigo_SAP', 'Componente', 'Q'].

    Returns:
        pd.DataFrame: dfActiveCode con columnas de stock agregado.
    """
    
    dfConsInvPortalUniconSAP["Stock"] = pd.to_numeric(dfConsInvPortalUniconSAP["Stock"], errors="coerce")
    dfConsInvPortalUniconSAP = dfConsInvPortalUniconSAP[~((dfConsInvPortalUniconSAP["SAP"] == "A18110007355") & (dfConsInvPortalUniconSAP["Almacen"] == "300"))]
    
    dfBOMfinal["Q"] = pd.to_numeric(dfBOMfinal["Q"], errors="coerce")

    # Almacenes considerados como 'no disponible'
    almacenes_no_disponibles = {'104', '111', '112', '501'}

    # Agrupación para stock real
    stock_por_sap = dfConsInvPortalUniconSAP.groupby("SAP")["Stock"].agg([
        lambda x: x[dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum(),
        lambda x: x[~dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum()
    ]).reset_index()

    stock_por_sap.columns = ["Codigo_SAP", "Stock no Disponible", "Stock Disponible"]

    # Crear copia para stock equivalente en componentes
    stock_equivalente = stock_por_sap.merge(dfBOMfinal, left_on="Codigo_SAP", right_on="Codigo_SAP", how="inner")

    # Calcular stock equivalente por componente
    stock_equivalente["Stock no Disponible-C"] = stock_equivalente["Stock no Disponible"] * stock_equivalente["Q"]
    stock_equivalente["Stock Disponible-C"] = stock_equivalente["Stock Disponible"] * stock_equivalente["Q"]

    # Agrupar por Componente
    stock_componentes = stock_equivalente.groupby("Componente")[["Stock no Disponible-C", "Stock Disponible-C"]].sum().reset_index()

    # Renombrar para merge
    stock_componentes.rename(columns={"Componente": "Codigo_SAP"}, inplace=True)

    # Unir ambos cálculos al dfActiveCode
    dfActiveCode = dfActiveCode.merge(stock_por_sap, on="Codigo_SAP", how="left")
    dfActiveCode = dfActiveCode.merge(stock_componentes, on="Codigo_SAP", how="left")

    # Rellenar valores faltantes con 0
    dfActiveCode[["Stock no Disponible", "Stock Disponible", "Stock no Disponible-C", "Stock Disponible-C"]] = \
        dfActiveCode[["Stock no Disponible", "Stock Disponible", "Stock no Disponible-C", "Stock Disponible-C"]].fillna(0)

    return dfActiveCode






def agregar_stock_disponibleFR(dfActiveCode, dfConsInvPortalUniconSAP):
    """
    Agrega las columnas 'Stock Disponible' y 'Stock no Disponible' a dfActiveCode.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        dfConsInvPortalUniconSAP (pd.DataFrame): DataFrame con la información de stock y almacenes.

    Returns:
        pd.DataFrame: dfActiveCode con las nuevas columnas agregadas.
    """
    # Definir los almacenes considerados como 'no disponible'
    almacenes_no_disponibles = {'104', '111', '112', '113'}

    # Agrupar por SAP y calcular los valores
    stock_por_sap = dfConsInvPortalUniconSAP.groupby("GET")["Stock"].agg([
        lambda x: x[dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum(),
        lambda x: x[~dfConsInvPortalUniconSAP.loc[x.index, "Almacen"].isin(almacenes_no_disponibles)].sum()
    ]).reset_index()

    # Renombrar columnas
    stock_por_sap.columns = ["GET", "Stock no Disponible", "Stock Disponible"]

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(stock_por_sap, left_on="Codigo_GET", right_on="GET", how="left").fillna(0)

    return dfActiveCode

def agregar_stock_disponibleMNG(dfActiveCode, dfConsInvPortalUniconSAP):

    """
    Agrega las columnas 'Stock Disponible', 'Stock no Disponible' y 'StockCallao' a dfActiveCode.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        dfConsInvPortalUniconSAP (pd.DataFrame): DataFrame con la información de stock y almacenes.

    Returns:
        pd.DataFrame: dfActiveCode con las nuevas columnas agregadas.
    """
    # Definir los almacenes considerados como 'no disponible' y los de Callao y CD
    almacenes_no_disponibles = {'104', '108', '111', '112', '501', '502'}
    callaoycd = {'100', '101', '102', '105', '106', '110', '114', '400', '495', '395', '500', '503', '199'}  # Convertido a enteros
    sucursales = {'210', '220', '230', '232', '240', '241'}
    chinalco = {'360'}  # Convertido a enteros
    kmmp = {'200'}  # Convertido a enteros
    apm = {'APM-01'}
    cosapi = {'330'}
    bambas = {'300'}
    tiamaria = {'385'} 
    

    # Filtrar el DataFrame en tres categorías
    df_no_disponible = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(almacenes_no_disponibles)]
    df_disponible = dfConsInvPortalUniconSAP[~dfConsInvPortalUniconSAP["Almacen"].isin(almacenes_no_disponibles)]
    df_callao = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(callaoycd)]
    df_sucursales = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(sucursales)]
    df_chinalco = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(chinalco)]
    df_kmmp = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(kmmp)]
    df_apm = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(apm)]
    df_cosapi = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(cosapi)]
    df_bambas = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(bambas)]
    df_tiamaria = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(tiamaria)]  

    
    # Agrupar y sumar por SAP
    stock_no_disponible = df_no_disponible.groupby("SAP")["Stock"].sum().rename("Stock no Disponible")
    stock_disponible = df_disponible.groupby("SAP")["Stock"].sum().rename("Stock Disponible")
    stock_callao = df_callao.groupby("SAP")["Stock"].sum().rename("StockCallao")
    stock_sucursales = df_sucursales.groupby("SAP")["Stock"].sum().rename("StockSucursales")
    stock_chinalco = df_chinalco.groupby("SAP")["Stock"].sum().rename("StockCHIN")
    stock_kmmp = df_kmmp.groupby("SAP")["Stock"].sum().rename("StockKMMP")
    stock_apm = df_apm.groupby("SAP")["Stock"].sum().rename("StockAPM")
    stock_cosapi = df_cosapi.groupby("SAP")["Stock"].sum().rename("StockCOS")
    stock_bambas = df_bambas.groupby("SAP")["Stock"].sum().rename("StockBAMB")
    stock_tiamaria = df_tiamaria.groupby("SAP")["Stock"].sum().rename("StockTíaMaría")

    
    # Combinar en un solo DataFrame
    stock_por_sap = pd.concat([stock_no_disponible, stock_disponible, stock_callao,stock_sucursales,stock_chinalco, stock_kmmp, stock_apm, stock_cosapi, stock_bambas,stock_tiamaria], axis=1).reset_index()

    # Renombrar la columna 'SAP' para coincidir con 'Codigo_SAP' en dfActiveCode
    stock_por_sap.rename(columns={"SAP": "Codigo_SAP"}, inplace=True)

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(stock_por_sap, on="Codigo_SAP", how="left").fillna(0)

    return dfActiveCode,stock_por_sap


def agregar_stock_disponibleAST(dfActiveCode, dfConsInvPortalUniconSAP):

    """
    Agrega columnas de stock por almacén a dfActiveCode.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con 'Codigo_SAP'
        dfConsInvPortalUniconSAP (pd.DataFrame): DataFrame con columnas SAP, Almacen, Stock

    Returns:
        pd.DataFrame
    """
    dfConsInvPortalUniconSAP["Almacen"] = dfConsInvPortalUniconSAP["Almacen"].astype(str)
    callaoycd = {'100','101','102','105','106','110','114','113','400','495','395','500','502','503'}
    jic = {'800'}
    hb = {'801'}
    toromocho = {'802'}
    antamina = {'803'}

    df_callao = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(callaoycd)]
    df_jic = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(jic)]
    df_hb = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(hb)]
    df_toro = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(toromocho)]
    df_anta = dfConsInvPortalUniconSAP[dfConsInvPortalUniconSAP["Almacen"].isin(antamina)]
    print("Registros Toromocho:", len(df_toro))
    print("Registros Antamina:", len(df_anta))
    stock_callao = df_callao.groupby("SAP")["Stock"].sum().rename("StockCallao")
    stock_jic = df_jic.groupby("SAP")["Stock"].sum().rename("StockJIC")
    stock_hb = df_hb.groupby("SAP")["Stock"].sum().rename("StockHB")
    stock_por_sap = pd.concat([stock_callao, stock_jic, stock_hb], axis=1).reset_index()
    

    stock_por_sap.rename(columns={"SAP": "Codigo_SAP"}, inplace=True)

    dfActiveCode = dfActiveCode.merge(stock_por_sap, on="Codigo_SAP", how="left")
    columnas_base = ["StockCallao","StockJIC","StockHB"]
    dfActiveCode[columnas_base] = dfActiveCode[columnas_base].fillna(0)
    if df_toro.empty:
        dfActiveCode["StockToromocho"] = 0
    else:
        stock_toro = df_toro.groupby("SAP")["Stock"].sum()
        dfActiveCode = dfActiveCode.merge(
            stock_toro.rename("StockToromocho"),
            left_on="Codigo_SAP",
            right_index=True,
            how="left"
        ).fillna({"StockToromocho":0})

    if df_anta.empty:
        dfActiveCode["StockAntamina"] = 0
    else:
        stock_anta = df_anta.groupby("SAP")["Stock"].sum()
        dfActiveCode = dfActiveCode.merge(
            stock_anta.rename("StockAntamina"),
            left_on="Codigo_SAP",
            right_index=True,
            how="left"
        ).fillna({"StockAntamina":0})    

   # for c in columnas:
   #     if c not in dfActiveCode.columns:
   #         dfActiveCode[c] = 0

   # dfActiveCode[columnas] = dfActiveCode[columnas].fillna(0)

    return dfActiveCode


def agregar_categoria_rotacion(dfActiveCode, ResumenPrevKardexSAPsinAlmacen, reference_period):
    """
    Agrega la columna 'Categoria Rot' a dfActiveCode basada en la cantidad de meses con consumo > 0.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        ResumenPrevKardexSAPsinAlmacen (pd.DataFrame): DataFrame con el historial de consumo.
        reference_period (pd.Timestamp): Fecha de referencia para calcular los últimos 12 meses.

    Returns:
        pd.DataFrame: dfActiveCode con la nueva columna 'Categoria Rot'.
    """
    # Convertir month_year a tipo datetime
    ResumenPrevKardexSAPsinAlmacen['month_year'] = pd.to_datetime(ResumenPrevKardexSAPsinAlmacen['month_year'])

    # Definir el rango de fechas de los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar datos dentro del rango de 12 meses
    df_filtrado = ResumenPrevKardexSAPsinAlmacen[
        (ResumenPrevKardexSAPsinAlmacen['month_year'] >= start_date) &
        (ResumenPrevKardexSAPsinAlmacen['month_year'] <= reference_period)
    ]

    # Contar la cantidad de meses con consumo > 0 por Código SAP
    meses_con_consumo = df_filtrado[df_filtrado['Consumo Total'] > 0].groupby('SAP')['month_year'].count().reset_index()
    ##
    meses_con_consumo.rename(columns={'SAP': 'Codigo_SAP','month_year': 'N° Meses'}, inplace=True)
    ##
    # Asignar la categoría de rotación
    def clasificar_rotacion(meses):
        if meses == 0:
            return "NULA"
        elif meses <= 2:
            return "BAJA"
        elif meses <= 6:
            return "MEDIA"
        else:
            return "ALTA"
    meses_con_consumo['Categoria Rot'] = meses_con_consumo['N° Meses'].apply(clasificar_rotacion)
    #meses_con_consumo['Categoria Rot'] = meses_con_consumo['month_year'].apply(clasificar_rotacion)
    #meses_con_consumo.rename(columns={'SAP': 'Codigo_SAP'}, inplace=True)

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(meses_con_consumo[['Codigo_SAP', 'Categoria Rot','N° Meses' ]], on='Codigo_SAP', how='left')

    # Llenar NaN con "NULA" para los códigos sin consumo registrado
    dfActiveCode['Categoria Rot'].fillna("NULA", inplace=True)
    dfActiveCode['N° Meses'].fillna(0, inplace=True)##

    return dfActiveCode

def agregar_categoria_grupal(dfActiveCode):

    prioridad = {
        "NULA": 0,
        "BAJA": 1,
        "MEDIA": 2,
        "ALTA": 3
    }

    inverso = {v: k for k, v in prioridad.items()}

    # calcular rotación mayor del grupo
    dfActiveCode["Categoria Grupal"] = (
        dfActiveCode["Categoria Rot"]
        .map(prioridad)
        .groupby(dfActiveCode["GRUPOS ADAN M"])
        .transform("max")
        .map(inverso)
    )

    return dfActiveCode

def agregar_categoria_rotacion_comp(dfActiveCode, df_componente_consumo, reference_period):
    """
    Agrega la columna 'Categoria Rot' a dfActiveCode basada en la cantidad de meses con consumo > 0.

    Args:
        dfActiveCode (pd.DataFrame): DataFrame principal con la columna 'Codigo_SAP'.
        ResumenPrevKardexSAPsinAlmacen (pd.DataFrame): DataFrame con el historial de consumo.
        reference_period (pd.Timestamp): Fecha de referencia para calcular los últimos 12 meses.

    Returns:
        pd.DataFrame: dfActiveCode con la nueva columna 'Categoria Rot'.
    """
    # Convertir month_year a tipo datetime
    df_componente_consumo['month_year'] = pd.to_datetime(df_componente_consumo['month_year'])

    # Definir el rango de fechas de los últimos 12 meses
    start_date = reference_period - pd.DateOffset(months=11)

    # Filtrar datos dentro del rango de 12 meses
    df_filtrado = df_componente_consumo[
        (df_componente_consumo['month_year'] >= start_date) &
        (df_componente_consumo['month_year'] <= reference_period)
    ]

    # Contar la cantidad de meses con consumo > 0 por Código SAP
    meses_con_consumo = df_filtrado[df_filtrado['Consumo Total'] > 0].groupby('SAP')['month_year'].count().reset_index()

    # Asignar la categoría de rotación
    def clasificar_rotacion(meses):
        if meses == 0:
            return "NULA"
        elif meses <= 2:
            return "BAJA"
        elif meses <= 6:
            return "MEDIA"
        else:
            return "ALTA"

    meses_con_consumo['Categoria Rot Comp'] = meses_con_consumo['month_year'].apply(clasificar_rotacion)
    meses_con_consumo.rename(columns={'SAP': 'Codigo_SAP'}, inplace=True)

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(meses_con_consumo[['Codigo_SAP', 'Categoria Rot Comp']], on='Codigo_SAP', how='left')

    # Llenar NaN con "NULA" para los códigos sin consumo registrado
    dfActiveCode['Categoria Rot Comp'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_mesesconsumoHYD(dfActiveCode, ResumenPrevKardexSAPsinAlmacen, reference_period):

    # Convertir month_year a tipo datetime
    ResumenPrevKardexSAPsinAlmacen['month_year'] = pd.to_datetime(ResumenPrevKardexSAPsinAlmacen['month_year'])

    # Contar la cantidad de meses con consumo > 0 por Código SAP
    meses_con_consumo = ResumenPrevKardexSAPsinAlmacen[ResumenPrevKardexSAPsinAlmacen['Consumo Total'] > 0].groupby('SAP')['month_year'].count().reset_index()
    meses_con_consumo.rename(columns={'SAP': 'Codigo_SAP', 'month_year': 'Rotacion'}, inplace=True)

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(meses_con_consumo[['Codigo_SAP', 'Rotacion']], on='Codigo_SAP', how='left')

    # Llenar NaN con "NULA" para los códigos sin consumo registrado
    dfActiveCode['Rotacion'].fillna(0, inplace=True)

    return dfActiveCode


def agregar_coef_variacionHYD(dfActiveCode, ResumenPrevKardexSAPsinAlmacen, reference_period):
    
    # Asegurar que month_year sea datetime
    ResumenPrevKardexSAPsinAlmacen['month_year'] = pd.to_datetime(ResumenPrevKardexSAPsinAlmacen['month_year'])

    # Agrupar por SAP y calcular media y desviación estándar del consumo
    stats_consumo = ResumenPrevKardexSAPsinAlmacen.groupby('SAP')['Consumo Total'].agg(['mean', 'std']).reset_index()

    # Calcular el coeficiente de variación (CV)
    stats_consumo['Coef_Variacion'] = stats_consumo['std'] / stats_consumo['mean']

    # Reemplazar infinities y NaN con 0 (por ejemplo, cuando media = 0)
    stats_consumo['Coef_Variacion'].replace([np.inf, -np.inf], 100, inplace=True)
    stats_consumo['Coef_Variacion'] = stats_consumo['Coef_Variacion'].fillna(0)

    # Renombrar columna SAP para hacer merge
    stats_consumo.rename(columns={'SAP': 'Codigo_SAP'}, inplace=True)

    # Merge con dfActiveCode
    dfActiveCode = dfActiveCode.merge(stats_consumo[['Codigo_SAP', 'Coef_Variacion']], on='Codigo_SAP', how='left')

    return dfActiveCode

def calcular_tendencia_consumo(df_activos, df_consumo):

    # Asegurar tipos
    df_consumo['month_year'] = pd.to_datetime(df_consumo['month_year'])
    # Codificar tiempo como número entero (ej. meses desde el inicio)
    df_consumo['Mes'] = (df_consumo['month_year'].dt.year - df_consumo['month_year'].dt.year.min()) * 12 + df_consumo['month_year'].dt.month

    def ajustar_regresion(grupo):
        if len(grupo) < 2:
            return pd.Series({'Pendiente': 0})
        X = grupo[['Mes']]
        y = grupo['Consumo Total']
        modelo = LinearRegression()
        modelo.fit(X, y)
        return pd.Series({'Pendiente': modelo.coef_[0]})

    # Aplicar regresión por 'Codigo SAP'
    pendientes = df_consumo.groupby('SAP').apply(ajustar_regresion).reset_index()

    # Unir con códigos activos para mantener solo los activos
    resultado =pd.merge(df_activos, pendientes, left_on='Codigo_SAP', right_on='SAP', how='left')
    resultado = resultado.drop(columns=['SAP'])

    # Reemplazar NaN (sin data de consumo) con 0
    resultado['Pendiente'] = resultado['Pendiente'].fillna(0)
    
    return resultado

def asignar_grupo_y_jerarquia(dfActiveCode, dfBOMfinal):
    # Crear una lista de conexiones: [(A, B), (A, C), ...]
    conexiones = dfBOMfinal[["Codigo_SAP", "Componente"]].values.tolist()

    # Crear un grafo no dirigido
    G = nx.Graph()
    G.add_edges_from(conexiones)

    # Obtener los componentes conexos (subgrafos conectados)
    componentes_conexos = list(nx.connected_components(G))

    # Asignar un grupo a cada nodo
    nodo_grupo = {}
    for i, grupo in enumerate(componentes_conexos, start=1):
        for nodo in grupo:
            nodo_grupo[nodo] = f"G{i}"

    # Crear DataFrame con asignación de grupo
    df_grupo = pd.DataFrame.from_dict(nodo_grupo, orient="index", columns=["Grupo"])
    df_grupo.index.name = "Codigo_SAP"
    df_grupo.reset_index(inplace=True)

    # Identificar si un código es producto final (aparece en Codigo_SAP del BOM)
    codigos_producto = set(dfBOMfinal["Codigo_SAP"])
    df_grupo["EsProducto"] = df_grupo["Codigo_SAP"].apply(lambda x: 1 if x in codigos_producto else 0)

    # Unir con dfActiveCode
    dfActiveCode = dfActiveCode.merge(df_grupo, on="Codigo_SAP", how="left")

    # Para los que no tienen grupo (no están en BOM), se les puede asignar 'Gx'
    max_grupo = len(componentes_conexos)
    dfActiveCode["Grupo"] = dfActiveCode["Grupo"].fillna(
        dfActiveCode["Codigo_SAP"].apply(lambda x: f"G{max_grupo + 1}")
    )
    dfActiveCode["EsProducto"] = dfActiveCode["EsProducto"].fillna(1)  # Si no es componente, asumir producto

    # Ordenar por Grupo y EsProducto (productos primero)
    dfActiveCode = dfActiveCode.sort_values(by=["Grupo", "EsProducto"], ascending=[True, False]).reset_index(drop=True)
    dfActiveCode = dfActiveCode.drop(columns="EsProducto")
        
    return dfActiveCode


def agregar_coberturas_y_ajustes(dfActiveCode):
    # Inicializar las columnas con NaN
    dfActiveCode["Cobertura SS (m)"] = None
    dfActiveCode["Cobertura RP (m)"] = None

    # Definir condiciones para ALTA y MEDIA
    cond_alta = dfActiveCode["Categoria Rot Comp"] == "ALTA"
    cond_media = dfActiveCode["Categoria Rot Comp"] == "MEDIA"

    # Calcular cobertura SS
    ss_alta = dfActiveCode.loc[cond_alta, "SS_98_ltvar"]
    ss_media = dfActiveCode.loc[cond_media, "SS_95_ltvar"]
    consumo = dfActiveCode["Consumo Promedio Comp 12M"]

    dfActiveCode.loc[cond_alta, "Cobertura SS (m)"] = ss_alta / consumo
    dfActiveCode.loc[cond_media, "Cobertura SS (m)"] = ss_media / consumo

    # Calcular cobertura RP
    rp_alta = dfActiveCode.loc[cond_alta, "P98_ltvar"]
    rp_media = dfActiveCode.loc[cond_media, "P95_ltvar"]

    dfActiveCode.loc[cond_alta, "Cobertura RP (m)"] = rp_alta / consumo
    dfActiveCode.loc[cond_media, "Cobertura RP (m)"] = rp_media / consumo

    # Manejo de errores: si la división fue válida, sumar 1
    for col in ["Cobertura SS (m)", "Cobertura RP (m)"]:
        dfActiveCode[col] = pd.to_numeric(dfActiveCode[col], errors="coerce")  # Asegurar numérico
        dfActiveCode[col] = dfActiveCode[col].where(dfActiveCode[col].notna(), 0)
        dfActiveCode[col] = dfActiveCode[col] + dfActiveCode[col].gt(0).astype(int)  # sumar 1 si > 0

    # Calcular ajustes
    dfActiveCode["SS Ajustado"] = dfActiveCode["Cobertura SS (m)"] * dfActiveCode["Promedio_3m_Forecast-C"]
    dfActiveCode["RP Ajustado"] = dfActiveCode["Cobertura RP (m)"] * dfActiveCode["Promedio_3m_Forecast-C"]

    return dfActiveCode
"""
def agregar_compar_precios(dfActiveCode, dfprecloimpor, sociedad):

    # Tipos de cambio estándar
    tipo_cambio = {
        'SOL': 1,
        'USD': 3.8,
        'EUR': 4.1,
        'CAD': 2.8,
        'GBP': 4.8
    }

    # Filtrar ambos DataFrames por sociedad
    df_precio = dfprecloimpor[dfprecloimpor['Sociedad'] == sociedad].copy()

    # Asegurar columnas numéricas
    df_precio['Precio Unit Local'] = pd.to_numeric(df_precio['Precio Unit Local'], errors='coerce')
    df_precio['Precio Unit Internacional'] = pd.to_numeric(df_precio['Precio Unit Internacional'], errors='coerce')

    # Hacer merge por ItemCode y Sociedad
    df_precio = df_precio.rename(columns={"ItemCode": "Codigo_SAP"})
    df_merge = dfActiveCode.merge(df_precio, on=['Codigo_SAP'], how='left')

    # Conversión de moneda a SOL
    df_merge['Cambio Local'] = df_merge['Moneda Local'].map(tipo_cambio)
    df_merge['Cambio Impor'] = df_merge['Moneda Internacional'].map(tipo_cambio)

    df_merge['Local_SOL'] = df_merge['Precio Unit Local'] * df_merge['Cambio Local']
    df_merge['Impor_SOL'] = df_merge['Precio Unit Internacional'] * df_merge['Cambio Impor']*1.07

    # Formateo de fechas
    def format_fecha(fecha):
        try:
            return pd.to_datetime(fecha).strftime('%m-%Y')
        except:
            return None

    df_merge['Fecha Local Fmt'] = df_merge['Fecha Local'].apply(format_fecha)
    df_merge['Fecha Impor Fmt'] = df_merge['Fecha Internacional'].apply(format_fecha)

    # Función para armar la cadena final por fila
    def construir_comparacion(row):
        loc_precio = row['Precio Unit Local']
        imp_precio = row['Precio Unit Internacional']
        loc_sol = row['Local_SOL']
        imp_sol = row['Impor_SOL']
        loc_fecha = row['Fecha Local Fmt']
        imp_fecha = row['Fecha Impor Fmt']
        loc_prov = row['Proveedor Local']
        imp_prov = row['Proveedor Internacional']
        mon_loc = row['Moneda Local']
        mon_imp = row['Moneda Internacional']

        # Formatos individuales
        loc_str = "No existe precio local"
        if pd.notna(loc_sol) and loc_fecha:
            loc_str = f"{loc_prov} ({loc_fecha} - {loc_precio:.2f} {mon_loc})"

        imp_str = "No existe precio importado"
        if pd.notna(imp_sol) and imp_fecha:
            imp_str = f"{imp_prov} ({imp_fecha} - {imp_precio:.2f} {mon_imp})"

        # Comparación
        if pd.notna(loc_sol) and pd.notna(imp_sol):
            variacion = (imp_sol - loc_sol) / loc_sol
            return f"{round(variacion * 100)}% // {loc_str} // {imp_str}"
        elif pd.isna(loc_sol) and pd.notna(imp_sol):
            return f"No existe precio local // {imp_str}"
        elif pd.notna(loc_sol) and pd.isna(imp_sol):
            return f"{loc_str} // No existe precio importado"
        else:
            return "No existe precios local e importado"

    # Aplicar lógica a cada fila (más rápido que for)
    df_merge['Comparacion Precio'] = df_merge.apply(construir_comparacion, axis=1)

    # Asignar al DataFrame original
#     dfActiveCode = dfActiveCode.copy()
#     dfActiveCode.loc['Comparacion Precio'] = df_merge['Comparacion Precio'].values
    
    dfActiveCode = dfActiveCode.merge(
    df_merge[['Codigo_SAP', 'Comparacion Precio']],
    on=['Codigo_SAP'],
    how='left')

    return dfActiveCode
"""

def agregar_compar_precios(dfActiveCode, dfprecloimpor, sociedad):

    # Tipos de cambio estándar
    tipo_cambio = {
        'SOL': 1,
        'USD': 3.8,
        'EUR': 4.1,
        'CAD': 2.8,
        'GBP': 4.8
    }

    # Filtrar ambos DataFrames por sociedad
    df_precio = dfprecloimpor[dfprecloimpor['Sociedad'] == sociedad].copy()

    # Asegurar columnas numéricas
    df_precio['Precio Unit Local'] = pd.to_numeric(df_precio['Precio Unit Local'], errors='coerce')
    df_precio['Precio Unit Internacional'] = pd.to_numeric(df_precio['Precio Unit Internacional'], errors='coerce')

    # Hacer merge por ItemCode y Sociedad
    df_precio = df_precio.rename(columns={"ItemCode": "Codigo_SAP"})
    df_merge = dfActiveCode.merge(df_precio, on=['Codigo_SAP'], how='left')

    # Conversión de moneda a SOL
    df_merge['Cambio Local'] = df_merge['Moneda Local'].map(tipo_cambio)
    df_merge['Cambio Impor'] = df_merge['Moneda Internacional'].map(tipo_cambio)

    df_merge['Local_SOL'] = df_merge['Precio Unit Local'] * df_merge['Cambio Local']
    df_merge['Impor_SOL'] = df_merge['Precio Unit Internacional'] * df_merge['Cambio Impor']*1.07
    
    df_merge['Variacion_Precio'] = np.where(
    df_merge['Local_SOL'].notna() &
    df_merge['Impor_SOL'].notna() &
    (df_merge['Local_SOL'] > 0),
    ((df_merge['Impor_SOL'] - df_merge['Local_SOL']) / df_merge['Local_SOL']) * 100,
    np.nan
    )

    # Formateo de fechas
    def format_fecha(fecha):
        try:
            return pd.to_datetime(fecha).strftime('%m-%Y')
        except:
            return None

    df_merge['Fecha Local Fmt'] = df_merge['Fecha Local'].apply(format_fecha)
    df_merge['Fecha Impor Fmt'] = df_merge['Fecha Internacional'].apply(format_fecha)

    # Función para armar la cadena final por fila
    def construir_comparacion(row):
        loc_precio = row['Precio Unit Local']
        imp_precio = row['Precio Unit Internacional']
        loc_sol = row['Local_SOL']
        imp_sol = row['Impor_SOL']
        loc_fecha = row['Fecha Local Fmt']
        imp_fecha = row['Fecha Impor Fmt']
        loc_prov = row['Proveedor Local']
        imp_prov = row['Proveedor Internacional']
        mon_loc = row['Moneda Local']
        mon_imp = row['Moneda Internacional']

        # Formatos individuales
        loc_str = "No existe precio local"
        if pd.notna(loc_sol) and loc_fecha:
            loc_str = f"{loc_prov} ({loc_fecha} - {loc_precio:.2f} {mon_loc})"

        imp_str = "No existe precio importado"
        if pd.notna(imp_sol) and imp_fecha:
            imp_str = f"{imp_prov} ({imp_fecha} - {imp_precio:.2f} {mon_imp})"

        # Comparación
        if pd.notna(loc_sol) and pd.notna(imp_sol):
            var = row['Variacion_Precio']
            return f"{round(var)}% // {loc_str} // {imp_str}"
        elif pd.isna(loc_sol) and pd.notna(imp_sol):
            return f"No existe precio local // {imp_str}"
        elif pd.notna(loc_sol) and pd.isna(imp_sol):
            return f"{loc_str} // No existe precio importado"
        else:
            return "No existe precios local e importado"

    # Aplicar lógica a cada fila (más rápido que for)
    df_merge['Comparacion Precio'] = df_merge.apply(construir_comparacion, axis=1)

    # Asignar al DataFrame original
#     dfActiveCode = dfActiveCode.copy()
#     dfActiveCode.loc['Comparacion Precio'] = df_merge['Comparacion Precio'].values
    
    dfActiveCode = dfActiveCode.merge(
    df_merge[['Codigo_SAP', 'Comparacion Precio', 'Variacion_Precio']],
    on='Codigo_SAP',
    how='left'
    )

    return dfActiveCode

def ajustar_clasificacion(df_setting):

    prioridad = {
        "Estrategico": 1,
        "Revision Periodica": 2,
        "Gestion Simplificada": 3,
        "Compra Calzada": 4
    }

    df = df_setting.copy()

    # Asegurar tipo string
    df["Producto Final"] = df["Producto Final"].fillna("").astype(str)

    # Detectar inicio de un nuevo grupo:
    # cuando Producto Final está vacío, pero la fila anterior no lo estaba
    df["nuevo_grupo"] = (df["Producto Final"] == "") & (df["Producto Final"].shift(1) != "")

    # Crear ID de grupo acumulado
    df["Grupo"] = df["nuevo_grupo"].cumsum()

    # Inicialmente la clasificación ajustada es igual a la original
    df["Clasificacion Ajustada"] = df["Clasificación"]

    # Ajustar clasificación por grupo
    for grupo, data in df.groupby("Grupo"):

        clasifs = data["Clasificación"].dropna()

        if clasifs.empty:
            continue

        # Convertir a prioridad numérica
        clasifs_map = clasifs.map(prioridad)

        # Tomar la clasificación más crítica (menor número)
        mejor = clasifs_map.min()

        clasificacion_final = [k for k, v in prioridad.items() if v == mejor][0]

        # Asignar a todo el grupo
        df.loc[data.index, "Clasificacion Ajustada"] = clasificacion_final

    return df.drop(columns=["nuevo_grupo", "Grupo"])

def revertir_reporte_anterior(df_pivot):
    columnas_fijas = [
        'SAP_Origen',
        'SAP',
        'Codigo_GET',
        'Descripcion',
        'UM',
        'Metodo',
        'LT_final',
        'Precision_elegido',
        'Nombre_Cliente'
    ]
    
    # Detectar columnas de meses
    columnas_meses = [c for c in df_pivot.columns if c not in columnas_fijas]
    
    # Volver de formato ancho a largo
    reporte_forecast = df_pivot.melt(
        id_vars=['Metodo', 'SAP', 'SAP_Origen'],
        value_vars=columnas_meses,
        var_name='Fecha',
        value_name='Forecast'
    )
    
    # Convertir Fecha nuevamente a datetime
    reporte_forecast['Fecha'] = pd.to_datetime(
        reporte_forecast['Fecha'],
        format='%b-%y'
    )
    
    # Eliminar filas vacías o en cero si deseas
    reporte_forecast = reporte_forecast[
        reporte_forecast['Forecast'].notna()
    ]
    
    # Ordenar
    reporte_forecast = reporte_forecast.sort_values(
        ['SAP', 'Fecha']
    ).reset_index(drop=True)
    
    reporte_forecast = reporte_forecast[
    ["SAP_Origen", "SAP", "Fecha", "Forecast"]
    ]   
    return reporte_forecast

def revertir_reporte_anterior_MGN(df_pivot):
    
    #df_pivot= df_forecast
    columnas_fijas = [
        'SAP_Origen',
        'SAP',
    ]
    
    # Detectar columnas de meses
    columnas_meses = [c for c in df_pivot.columns if c not in columnas_fijas]
    
    # Volver de formato ancho a largo
    reporte_forecast = df_pivot.melt(
        id_vars=['SAP', 'SAP_Origen'],
        value_vars=columnas_meses,
        var_name='Fecha',
        value_name='Forecast'
    )
    
    # Convertir Fecha nuevamente a datetime
    reporte_forecast['Fecha'] = pd.to_datetime(
        reporte_forecast['Fecha'],
        format='%b-%y'
    )
    
    # Eliminar filas vacías o en cero si deseas
    reporte_forecast = reporte_forecast[
        reporte_forecast['Forecast'].notna()
    ]
    
    # Ordenar
    reporte_forecast = reporte_forecast.sort_values(
        ['SAP', 'Fecha']
    ).reset_index(drop=True)
    
    reporte_forecast = reporte_forecast[
    ["SAP_Origen", "SAP", "Fecha", "Forecast"]
    ]   
    return reporte_forecast


def unidades_forecast(df_resultado_ltvar, df_forecast, columna_dinamica, primer_dia_mes_actual):

    primer_dia_mes_actual = primer_dia_mes_actual.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    # ===========================
    # TIPOS DE DATOS
    # ===========================
    df_forecast["Fecha"] = pd.to_datetime(
        df_forecast["Fecha"],
        errors="coerce"
    )

    df_forecast["Forecast"] = pd.to_numeric(
        df_forecast["Forecast"],
        errors="coerce"
    )

    # ===========================
    # AJUSTAR FORECAST MES ACTUAL
    # ===========================
    df_forecast = df_forecast.merge(
        df_resultado_ltvar[['Codigo_SAP', 'Consumo Mes Actual']],
        left_on='SAP_Origen',
        right_on='Codigo_SAP',
        how='left'
    )

    mask = df_forecast["Fecha"] == primer_dia_mes_actual

    df_forecast.loc[mask, "Forecast"] = np.where(
        df_forecast.loc[mask, "Consumo Mes Actual"] <
        df_forecast.loc[mask, "Forecast"],
        df_forecast.loc[mask, "Forecast"] -
        df_forecast.loc[mask, "Consumo Mes Actual"],
        0
    )

    df_forecast = df_forecast.drop(columns="Codigo_SAP")

    # ===========================
    # ORDENAR
    # ===========================
    df_forecast = df_forecast.sort_values(
        ["SAP_Origen", "Fecha"]
    )

    # ===========================
    # FRACCIÓN RESTANTE DEL MES
    # ===========================
    hoy = datetime.today()

    dias_mes = monthrange(
        hoy.year,
        hoy.month
    )[1]

    fraccion_mes_actual = 1 - (hoy.day / dias_mes)

    resultados = []

    # ===========================
    # RECORRER CADA SAP
    # ===========================
    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        cobertura = df_resultado_ltvar.loc[
            df_resultado_ltvar["Codigo_SAP"] == sap,
            columna_dinamica
        ]

        if cobertura.empty or pd.isna(cobertura.values[0]):
            resultados.append((sap, np.nan))
            continue

        cobertura = cobertura.values[0]

        unidades = 0

        primer_mes = True

        for row in grupo.itertuples():

            forecast = row.Forecast

            if pd.isna(forecast):
                continue

            # ===========================
            # PRIMER MES
            # ===========================
            if primer_mes:

                # Toda la cobertura está dentro del primer mes
                if cobertura <= fraccion_mes_actual:

                    if fraccion_mes_actual > 0:
                        unidades += (
                            cobertura /
                            fraccion_mes_actual
                        ) * forecast

                    cobertura = 0
                    break

                # La cobertura continúa a los siguientes meses
                else:

                    # El forecast YA representa únicamente
                    # la demanda restante del mes
                    unidades += forecast

                    cobertura -= fraccion_mes_actual

                    primer_mes = False

                    if cobertura <= 0:
                        break

            # ===========================
            # MESES SIGUIENTES
            # ===========================
            else:

                if cobertura >= 1:

                    unidades += forecast
                    cobertura -= 1

                else:

                    unidades += forecast * cobertura
                    cobertura = 0
                    break

        resultados.append((sap, unidades))

    # ===========================
    # RESULTADO
    # ===========================
    df_unidades = pd.DataFrame(
        resultados,
        columns=[
            "Codigo_SAP",
            "Unidades forecast"
        ]
    )

    df_resultado_ltvar = df_resultado_ltvar.merge(
        df_unidades,
        on="Codigo_SAP",
        how="left"
    )

    return df_resultado_ltvar

def unidades_forecast2(df_resultado_ltvar,df_forecast,columna_dinamica,primer_dia_mes_actual):
    ##Obtiene Unidades forecast Ajustado de "Cobertura (Punto de Reorden)" cuantas unidades representaran en forescast pasado el lead time
    #columna_dinamica="Cobertura (Punto de Reorden)"
    df_evaluar = df_resultado_ltvar[ ["Codigo_SAP", columna_dinamica, "Promedio_LT"] ].copy()

    df_evaluar[columna_dinamica] = pd.to_numeric( df_evaluar[columna_dinamica],errors="coerce" )
    df_evaluar["Promedio_LT"] = pd.to_numeric( df_evaluar["Promedio_LT"], errors="coerce" )
    df_evaluar = df_evaluar[ (df_evaluar[columna_dinamica].notna()) &  (df_evaluar[columna_dinamica] > 0) & (df_evaluar["Promedio_LT"].notna()) ]
    df_forecast = df_forecast.copy()
    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"], errors="coerce" )
    df_forecast["Forecast"] = pd.to_numeric( df_forecast["Forecast"], errors="coerce"  )

    df_forecast = pd.merge(df_forecast,df_resultado_ltvar[["Codigo_SAP", "Consumo Mes Actual"]],left_on="SAP_Origen",right_on="Codigo_SAP", how="left")
    df_forecast.drop(columns=["Codigo_SAP"], inplace=True)


    mask = (df_forecast["Fecha"]== primer_dia_mes_actual )

    df_forecast.loc[mask, "Forecast"] = np.where(
        df_forecast.loc[mask, "Consumo Mes Actual"]
        < df_forecast.loc[mask, "Forecast"],

        df_forecast.loc[mask, "Forecast"]
        - df_forecast.loc[mask, "Consumo Mes Actual"],

        0
    )

    df_forecast = df_forecast.sort_values( ["SAP_Origen", "Fecha"] )
    df_forecast["mes_idx"] = (df_forecast.groupby("SAP_Origen") .cumcount())

    df = df_forecast.merge(df_evaluar,left_on="SAP_Origen", right_on="Codigo_SAP", how="left")
    df["inicio"] = df["Promedio_LT"]
    df["fin"] = (df["Promedio_LT"] + df[columna_dinamica] )

    df["peso"] = np.clip(np.minimum(df["mes_idx"] + 1, df["fin"]) - np.maximum( df["mes_idx"], df["inicio"]), 0, 1 )
    df["forecast_ponderado"] = ( df["Forecast"]* df["peso"] )

    resultado = (df.groupby("Codigo_SAP")[ "forecast_ponderado"].sum().reset_index())

    resultado.rename( columns={ "forecast_ponderado":"Unidades forecast Ajustado"}, inplace=True )
    df_resultado_ltvar = ( df_resultado_ltvar.merge( resultado, on="Codigo_SAP", how="left" ))

    return df_resultado_ltvar

def extender_forecast(df_forecast, df_evaluar, columna_dinamica):

    df_forecast = df_forecast.copy()

    # Horizonte máximo requerido
    max_horizonte = math.ceil(
        (df_evaluar["Promedio_LT"] + df_evaluar[columna_dinamica]+1).max()
    )

    nuevas_filas = []

    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        grupo = grupo.sort_values("Fecha")

        meses_existentes = len(grupo)

        if meses_existentes >= max_horizonte:
            continue

        promedio = grupo["Forecast"].mean()

        ultima_fecha = grupo["Fecha"].max()

        meses_faltantes = max_horizonte - meses_existentes

        for i in range(1, meses_faltantes + 1):

            nuevas_filas.append({
                "SAP_Origen": sap,
                "Fecha": ultima_fecha + DateOffset(months=i),
                "Forecast": promedio
            })

    if nuevas_filas:
        df_forecast = pd.concat(
            [df_forecast, pd.DataFrame(nuevas_filas)],
            ignore_index=True
        )

    return df_forecast

def unidades_despues_LT(df_resultado_ltvar, df_forecast,columna_dinamica):
    
    # --- Coberturas ---    
    #columna_dinamica="Cobertura (Punto de Reorden)"
    columnas = ["Codigo_SAP", "Promedio_LT"]

    if columna_dinamica != "Promedio_LT":
        columnas.append(columna_dinamica)
    
    df_evaluar = df_resultado_ltvar[columnas].copy()
    
    df_evaluar[columna_dinamica] = pd.to_numeric(
        df_evaluar[columna_dinamica], errors="coerce"
    ).round(1)
    
    df_evaluar["Promedio_LT"] = pd.to_numeric(
        df_evaluar["Promedio_LT"], errors="coerce"
    ).round(1)
    
    df_evaluar = df_evaluar[
    df_evaluar["Promedio_LT"].notna() &
    np.isfinite(df_evaluar["Promedio_LT"]) &
    (df_evaluar["Promedio_LT"] != 0) &
    df_evaluar[columna_dinamica].notna() &
    np.isfinite(df_evaluar[columna_dinamica]) &
    (df_evaluar[columna_dinamica] > 0)
    ]
    
    # --- Forecast ---
    df_forecast = df_forecast.copy()
    
    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"], errors="coerce")
    df_forecast["Forecast"] = pd.to_numeric(df_forecast["Forecast"], errors="coerce")   
    
    df_forecast = df_forecast.sort_values(["SAP_Origen", "Fecha"]) 
    # Extender forecast si no alcanza el horizonte solicitado
    df_forecast = extender_forecast(
        df_forecast,
        df_evaluar,
        columna_dinamica
    )
    
    #Fraccion restante del mes actual
    hoy = datetime.today()
    dias_mes = calendar.monthrange(hoy.year, hoy.month)[1]
    fraccion_restante = (dias_mes - hoy.day) / dias_mes
    
    # índice de mes por SAP
    df_forecast["mes_idx"] = df_forecast.groupby("SAP_Origen").cumcount()
    
    # Inicio del intervalo temporal de cada mes
    df_forecast["inicio_mes"] = np.where(
        df_forecast["mes_idx"] == 0,
        0,
        fraccion_restante + (df_forecast["mes_idx"] - 1)
    )
    
    # Fin del intervalo temporal de cada mes
    df_forecast["fin_mes"] = np.where(
        df_forecast["mes_idx"] == 0,
        fraccion_restante,
        fraccion_restante + df_forecast["mes_idx"]
    )
    
    # --- merge ---
    df = df_forecast.merge(
        df_evaluar,
        left_on="SAP_Origen",
        right_on="Codigo_SAP",
        how="left"
    )
    
    df["inicio"] = df["Promedio_LT"]         
    df["fin"] = df["Promedio_LT"] + df[columna_dinamica]         
    
    df["peso"] = np.clip(
        np.minimum(df["fin_mes"], df["fin"]) -
        np.maximum(df["inicio_mes"], df["inicio"]),
        0,
        1
    )
    
    df["Unidades_despues_LT"] = df["Forecast"] * df["peso"]
    
    # --- suma final ---
    resultado = df.groupby("Codigo_SAP")["Unidades_despues_LT"].sum().reset_index()
 
    # --- merge final ---
    df_resultado_ltvar = df_resultado_ltvar.merge(resultado, on="Codigo_SAP", how="left")
     
    return df_resultado_ltvar


def sugerencia_compra_forecast(df_resultado_ltvar, df_forecast):

    df_resultado_ltvar = df_resultado_ltvar.copy()

    # Validar qué SAP existen en forecast
    condicion_forecast = df_resultado_ltvar["Codigo_SAP"].isin(df_forecast["SAP_Origen"])

    # Con forecast
    df_resultado_ltvar.loc[condicion_forecast, "Sugerencia de Compra"] = np.where(
        df_resultado_ltvar.loc[condicion_forecast, "Posicion de Inventario (PI)"]
        <= df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast"],

        np.minimum(
            df_resultado_ltvar.loc[condicion_forecast, "Unidades_segundo_LT"]
            + (
                df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast"]
                - df_resultado_ltvar.loc[condicion_forecast, "Posicion de Inventario (PI)"]
            ),

            df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast Ajustado"]
        ),

        np.nan
    )

    # Sin forecast
    df_resultado_ltvar.loc[~condicion_forecast, "Sugerencia de Compra"] = np.where(
        df_resultado_ltvar.loc[~condicion_forecast, "Posicion de Inventario (PI)"]
        <= df_resultado_ltvar.loc[~condicion_forecast, "Punto de reorden"],

        np.minimum(
            df_resultado_ltvar.loc[~condicion_forecast, "Consumo Promedio 12M"]
            * df_resultado_ltvar.loc[~condicion_forecast, "Promedio_LT"]

            + (
                df_resultado_ltvar.loc[~condicion_forecast, "Punto de reorden"]
                - df_resultado_ltvar.loc[~condicion_forecast, "Posicion de Inventario (PI)"]
            ),

            df_resultado_ltvar.loc[~condicion_forecast, "Punto de reorden"]
        ),

        np.nan
    )

    return df_resultado_ltvar

def sugerencia_compra_forecast_DOS(df_resultado_ltvar, df_forecast):

    df_resultado_ltvar = df_resultado_ltvar.copy()

    # Validar LT
    condicion_lt = (
        df_resultado_ltvar["Promedio_LT"].notna()
        & (df_resultado_ltvar["Promedio_LT"] != 0)
        & np.isfinite(df_resultado_ltvar["Promedio_LT"])
    )

    # Validar qué SAP existen en forecast
    condicion_forecast = (
        df_resultado_ltvar["Codigo_SAP"].isin(df_forecast["SAP_Origen"])
        & condicion_lt
    )

    # Validar SAP sin forecast
    condicion_sin_forecast = (
        ~df_resultado_ltvar["Codigo_SAP"].isin(df_forecast["SAP_Origen"])
        & condicion_lt
    )

    # -----------------------------
    # Con forecast
    # -----------------------------
    df_resultado_ltvar.loc[condicion_forecast, "Sugerencia de Compra"] = np.where(
        df_resultado_ltvar.loc[condicion_forecast, "Posicion de Inventario (PI)"]
        <= df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast"],

        np.minimum(
            df_resultado_ltvar.loc[condicion_forecast, "Unidades_segundo_LT"]
            + (
                df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast"]
                - df_resultado_ltvar.loc[condicion_forecast, "Posicion de Inventario (PI)"]
            ),

            df_resultado_ltvar.loc[condicion_forecast, "Punto de reorden - Forecast Ajustado"]
        ),

        np.nan
    )

    # -----------------------------
    # Sin forecast
    # -----------------------------
    df_resultado_ltvar.loc[condicion_sin_forecast, "Sugerencia de Compra"] = np.where(
        df_resultado_ltvar.loc[condicion_sin_forecast, "Posicion de Inventario (PI)"]
        <= df_resultado_ltvar.loc[condicion_sin_forecast, "Punto de reorden"],

        np.minimum(
            df_resultado_ltvar.loc[condicion_sin_forecast, "Consumo Promedio 12M"]
            * df_resultado_ltvar.loc[condicion_sin_forecast, "Promedio_LT"]

            + (
                df_resultado_ltvar.loc[condicion_sin_forecast, "Punto de reorden"]
                - df_resultado_ltvar.loc[condicion_sin_forecast, "Posicion de Inventario (PI)"]
            ),

            df_resultado_ltvar.loc[condicion_sin_forecast, "Punto de reorden"]
        ),

        np.nan
    )

    # LT inválido → sin sugerencia
    df_resultado_ltvar.loc[
        ~condicion_lt,
        "Sugerencia de Compra"
    ] = np.nan

    return df_resultado_ltvar


def cobertura_forecast(df_resultado_ltvar,df_forecast,primer_dia_mes_actual,columna_dinamica):
    
    #columna_dinamica = "Posicion de Inventario (PI)"
    primer_dia_mes_actual = primer_dia_mes_actual.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )
    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    #TIPOS DE DATOS#
    #df_forecast = df_forecast[df_forecast['SAP_Origen'] == 'A18110002915']
    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"],errors="coerce")
    df_forecast["Forecast"] = pd.to_numeric(df_forecast["Forecast"],errors="coerce")

    #MERGE CONSUMO MES ACTUAL#
    df_forecast = df_forecast.merge(
        df_resultado_ltvar[['Codigo_SAP', 'Consumo Mes Actual']],
        left_on='SAP_Origen',
        right_on='Codigo_SAP',
        how='left'
    )

    #AJUSTAR FORECAST MES ACTUAL# 
    mask = (df_forecast['Fecha']== primer_dia_mes_actual)
    df_forecast.loc[mask, 'Forecast'] = np.where(
        (df_forecast.loc[mask, 'Consumo Mes Actual']< df_forecast.loc[mask, 'Forecast']),
        (df_forecast.loc[mask, 'Forecast']- df_forecast.loc[mask, 'Consumo Mes Actual']),0)

    #ORDENAR#
    df_forecast = df_forecast.sort_values(
        ["SAP_Origen", "Fecha"]
    )
    

    resultados = []
    # FRACCIÓN DEL MES ACTUAL
    hoy = datetime.today()
    dias_mes = monthrange(
        hoy.year,
        hoy.month
    )[1]
    fraccion_mes_actual = (
        hoy.day / dias_mes
    )
    

    # SIMULACIÓN
    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        stock = df_resultado_ltvar.loc[(df_resultado_ltvar["Codigo_SAP"]== sap),columna_dinamica]

        # sin stock
        if (stock.empty or pd.isna(stock.values[0])):

            resultados.append((sap, np.nan))
            continue

        stock = stock.values[0]
        
        cobertura = 0
        # promedio mensual forecast
        forecast_promedio = (grupo["Forecast"].sum() / 12)
        primer_mes = True

        # RECORRER FORECAST
        for row in grupo.itertuples():

            consumo = row.Forecast

            # evitar NaN
            if pd.isna(consumo):
                continue

            # PRIMER MES
            if primer_mes:

                # si forecast restante = 0
                if consumo == 0:
                    cobertura += (fraccion_mes_actual)

                else:

                    # cobertura completa
                    if stock >= consumo:
                        stock = stock - consumo
                        cobertura += (fraccion_mes_actual)

                    # cobertura parcial
                    else:
                        cobertura += ((stock / consumo)* fraccion_mes_actual)
                        stock = 0
                        primer_mes = False
                        break
                primer_mes = False

            # MESES SIGUIENTES
            else:

                # evitar división por 0
                if consumo == 0:
                    cobertura += 1
                    continue

                # cobertura completa
                if stock >= consumo:
                    stock = stock - consumo
                    cobertura = (cobertura + 1)
                # cobertura parcial
                else:
                    cobertura += (stock / consumo)
                    stock = 0
                    break

        # STOCK SOBRANTE DESPUÉS DEL HORIZONTE
        if (stock > 0 and forecast_promedio > 0):

            cobertura_extra = (stock / forecast_promedio)
            cobertura += cobertura_extra

        resultados.append((sap, cobertura))

    # RESULTADO FINAL
    df_cobertura = pd.DataFrame(resultados,columns=["Codigo_SAP","Meses_Forecast"])
    df_resultado_ltvar = (
        df_resultado_ltvar.merge(
            df_cobertura,
            on="Codigo_SAP",
            how="left"
        )
    )

    return df_resultado_ltvar

def cobertura_forecast3( df_resultado_ltvar, df_forecast, primer_dia_mes_actual, columna_dinamica):
    """
    Convierte unidades forecast a meses de cobertura.

    Esta función es la inversa de unidades_forecast2().

    """
    primer_dia_mes_actual = primer_dia_mes_actual.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )
    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"], errors="coerce")
    df_forecast["Forecast"] = pd.to_numeric(df_forecast["Forecast"],errors="coerce")
    df_forecast = pd.merge(df_forecast,df_resultado_ltvar[["Codigo_SAP", "Consumo Mes Actual"]],left_on="SAP_Origen",right_on="Codigo_SAP",how="left" )
    mask = df_forecast["Fecha"] == primer_dia_mes_actual

    df_forecast.loc[mask, "Forecast"] = np.where(
        df_forecast.loc[mask, "Consumo Mes Actual"]
        < df_forecast.loc[mask, "Forecast"],

        df_forecast.loc[mask, "Forecast"]
        - df_forecast.loc[mask, "Consumo Mes Actual"],

        0
    )

    # ==========================================================
    # Generar índice secuencial de meses
    #
    # mes_idx:
    #   0 = mes actual
    #   1 = siguiente mes
    #   2 = siguiente
    # ==========================================================

    df_forecast = df_forecast.sort_values(["SAP_Origen", "Fecha"] )
    df_forecast["mes_idx"] = (df_forecast.groupby("SAP_Origen").cumcount() )
    resultados = []

    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        # ------------------------------------------------------
        # Unidades objetivo que queremos convertir a meses
        # ------------------------------------------------------

        stock_raw = df_resultado_ltvar.loc[df_resultado_ltvar["Codigo_SAP"] == sap, columna_dinamica]
        if stock_raw.empty or pd.isna(stock_raw.iloc[0]):
            resultados.append((sap, np.nan))
            continue
        unidades_objetivo = float(stock_raw.iloc[0])

        # ------------------------------------------------------
        # Lead Time
        # ------------------------------------------------------

        lt_raw = df_resultado_ltvar.loc[df_resultado_ltvar["Codigo_SAP"] == sap,"Promedio_LT"]

        lt = (
            0.0
            if lt_raw.empty or pd.isna(lt_raw.iloc[0])
            else float(lt_raw.iloc[0])
        )

        grupo = grupo.reset_index(drop=True)

        # Función auxiliar:

        def calcular_unidades(cobertura):

            inicio = lt
            fin = lt + cobertura

            total = 0.0

            for row in grupo.itertuples():

                forecast = row.Forecast

                if pd.isna(forecast) or forecast <= 0:
                    continue

                mes = row.mes_idx

                # ---------------------------------------------
                # Peso del mes dentro de la ventana
                #
                # MISMA fórmula que unidades_forecast2
                # ---------------------------------------------

                peso = np.clip(
                    min(mes + 1, fin)
                    - max(mes, inicio),
                    0,
                    1
                )

                total += forecast * peso

            return total

        # ======================================================
        # Búsqueda binaria
        #
        # Encontrar cobertura tal que:
        #
        # calcular_unidades(cobertura)
        #        =
        # unidades_objetivo
        # ======================================================

        low = 0.0
        high = 120.0  

        for _ in range(60):

            mid = (low + high) / 2

            unidades_mid = calcular_unidades(mid)

            if unidades_mid < unidades_objetivo:
                low = mid
            else:
                high = mid

        cobertura = (low + high) / 2

        resultados.append((sap, cobertura))

    # ==========================================================

    df_cobertura = pd.DataFrame(resultados,   columns=["Codigo_SAP", "Meses_Forecast"]  )
    df_resultado_ltvar = df_resultado_ltvar.merge(df_cobertura,on="Codigo_SAP",  how="left"   )

    return df_resultado_ltvar

def cobertura_forecast4(df_resultado_ltvar,df_forecast,  primer_dia_mes_actual, columna_dinamica):
# Esta función es la inversa de unidades_forecast2().
    primer_dia_mes_actual = primer_dia_mes_actual.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )
    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    df_forecast["Fecha"] = pd.to_datetime( df_forecast["Fecha"],  errors="coerce" )
    df_forecast["Forecast"] = pd.to_numeric( df_forecast["Forecast"], errors="coerce" )
    # Ajuste mes actual
    df_forecast = pd.merge(df_forecast,df_resultado_ltvar[ ["Codigo_SAP", "Consumo Mes Actual"]],  left_on="SAP_Origen",right_on="Codigo_SAP",   how="left")
    mask = df_forecast["Fecha"] == primer_dia_mes_actual

    df_forecast.loc[mask, "Forecast"] = np.where(
        df_forecast.loc[mask, "Consumo Mes Actual"]
        < df_forecast.loc[mask, "Forecast"],

        df_forecast.loc[mask, "Forecast"]
        - df_forecast.loc[mask, "Consumo Mes Actual"],

        0
    )

    df_forecast = df_forecast.sort_values(  ["SAP_Origen", "Fecha"]  )
    df_forecast["mes_idx"] = (   df_forecast.groupby("SAP_Origen")  .cumcount() )
    resultados = []
    for sap, grupo in df_forecast.groupby("SAP_Origen"):
        stock_raw = df_resultado_ltvar.loc[ df_resultado_ltvar["Codigo_SAP"] == sap,  columna_dinamica  ]
        if stock_raw.empty or pd.isna(stock_raw.iloc[0]):
            resultados.append((sap, np.nan))
            continue

        stock_objetivo = float(stock_raw.iloc[0])

        lt_raw = df_resultado_ltvar.loc[  df_resultado_ltvar["Codigo_SAP"] == sap, "Promedio_LT"]
        lt = (
            0.0
            if lt_raw.empty or pd.isna(lt_raw.iloc[0])
            else float(lt_raw.iloc[0])
        )

        grupo = grupo.reset_index(drop=True)

        acumulado = 0.0
        cobertura = 0.0

        for row in grupo.itertuples():

            forecast = row.Forecast

            if pd.isna(forecast) or forecast <= 0:
                continue

            mes = row.mes_idx

            peso_mes_completo = np.clip(
                min(mes + 1, lt + cobertura + 1)
                - max(mes, lt + cobertura),
                0,
                1
            )

            if peso_mes_completo <= 0:
                continue

            aporte_mes_completo = forecast * peso_mes_completo

            if acumulado + aporte_mes_completo <= stock_objetivo:

                acumulado += aporte_mes_completo
                cobertura += peso_mes_completo

            else:

                faltante = stock_objetivo - acumulado

                cobertura += faltante / forecast

                acumulado = stock_objetivo

                break

        resultados.append((sap, cobertura))

    df_cobertura = pd.DataFrame(resultados, columns=["Codigo_SAP", "Meses_Forecast"])
    df_resultado_ltvar = df_resultado_ltvar.merge(df_cobertura, on="Codigo_SAP", how="left")

    return df_resultado_ltvar

def cobertura_stockfisico_forecast_MGN(df_resultado_ltvar, df_forecast, primer_dia_mes_actual):

    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    # Tipo dato
    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"], errors="coerce")
    df_forecast["Forecast"] = pd.to_numeric(df_forecast["Forecast"], errors="coerce")

    # mes actual
    df_forecast = df_forecast.merge(
        df_resultado_ltvar[['Codigo_SAP', 'Consumo Mes Actual']], 
        left_on='SAP_Origen', 
        right_on='Codigo_SAP', 
        how='left'
    )

    # Ajuste mes actual
    mask = df_forecast['Fecha'] == primer_dia_mes_actual   
    df_forecast.loc[mask, 'Forecast'] = np.where(
        df_forecast.loc[mask, 'Consumo Mes Actual'] < df_forecast.loc[mask, 'Forecast'],
        df_forecast.loc[mask, 'Forecast'] - df_forecast.loc[mask, 'Consumo Mes Actual'],
        0
    )

    # Orden
    df_forecast = df_forecast.sort_values(["SAP_Origen", "Fecha"])
    resultados = []

    # Simulado
    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        stock = df_resultado_ltvar.loc[
            df_resultado_ltvar["Codigo_SAP"] == sap, 
            "Stock Disponible"
        ]
        if stock.empty or pd.isna(stock.values[0]):
            resultados.append((sap, np.nan))
            continue
        stock = stock.values[0]
        cobertura = 0
        for row in grupo.itertuples():
            consumo = row.Forecast
            if pd.isna(consumo) or consumo == 0:
                continue
            if stock >= consumo:
                stock -= consumo
                cobertura += 1
            else:
                cobertura += stock / consumo
                break
        resultados.append((sap, cobertura))

    df_cobertura = pd.DataFrame(resultados, columns=["Codigo_SAP", "Cobertura Stock Fisico - Forecast"])
    # merge final
    df_resultado_ltvar = df_resultado_ltvar.merge(
        df_cobertura, on="Codigo_SAP", how="left"
    )

    return df_resultado_ltvar

def clasificacion_comp(df_resultado_ltvar,dfMRP_filter,df_clasificacion):
   
    df = dfMRP_filter.copy() 
    df = df.reset_index().copy()
    # Asegurar string
    df["Producto Final"] = df["Producto Final"].astype(str)
    df["Componente"] = df["Componente"].astype(str)
    
    parent = {}
    
    def find(x):
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]
    
    def union(x, y):
        root_x = find(x)
        root_y = find(y)
        if root_x != root_y:
            parent[root_y] = root_x
    
    # Inicializar nodos (productos y componentes)
    nodos = set(df["Producto Final"]).union(set(df["Componente"]))
    for n in nodos:
        parent[n] = n
    
    # Unir producto con su componente
    for _, row in df.iterrows():
        union(row["Producto Final"], row["Componente"])
    
    # -------- CREAR GRUPOS --------
    grupo_map = {}
    grupo_id = 1
    
    for prod in df["Producto Final"].unique():
        root = find(prod)
        
        if root not in grupo_map:
            grupo_map[root] = grupo_id
            grupo_id += 1
    
    # Asignar grupo a cada producto
    df_grupos = pd.DataFrame({
        "SAP": df["Producto Final"].unique()
    })
    
    df_grupos["Grupo"] = df_grupos["SAP"].map(lambda x: grupo_map[find(x)])
    df_grupos["Tipo"] = "Producto Final"
    
    # -------- COMPONENTES --------
    df_comp = pd.DataFrame({
        "SAP": df["Componente"].unique()
    })
    
    df_comp["Grupo"] = df_comp["SAP"].map(lambda x: grupo_map[find(x)])
    df_comp["Tipo"] = "Componente"
    
    # -------- UNIR RESULTADO --------
    df_final = pd.concat([df_grupos, df_comp], ignore_index=True)
    
    # Orden opcional
    df_final = df_final.sort_values("Grupo").reset_index(drop=True)
    
    #Traemos clasificacion
    df_final = df_final.merge(df_clasificacion[['Codigo_SAP', 'Clasificación General']], 
                                      left_on='SAP', 
                                      right_on='Codigo_SAP', 
                                      how='left')
    df_final = df_final.drop(columns=["Codigo_SAP"])
    
    orden_categoria = {
    "Estrategico": 1,
    "Revision Periodica": 2,
    "Gestion Simplificada": 3,
    "Compra Calzada": 4
    }
    
    # Mapear prioridad numérica
    df_final["prio"] = df_final["Clasificación General"].map(orden_categoria)
    
    # Obtener la mejor prioridad por grupo
    mejor_prio = df_final.groupby("Grupo")["prio"].transform("min")
    
    # Mapear de vuelta a texto
    inv_map = {v: k for k, v in orden_categoria.items()}
    df_final["clasif_grupo"] = mejor_prio.map(inv_map)
    
    # Asignar SOLO a componentes
    df_final.loc[df_final["Tipo"] == "Componente", "Clasificación General"] = df_final.loc[
        df_final["Tipo"] == "Componente", "clasif_grupo"
    ]
    
    # Limpiar columnas auxiliares
    df_final.drop(columns=["prio", "clasif_grupo"], inplace=True)
    df_final = df_final[df_final['Tipo'] == 'Componente']
    df_final = df_final.rename(columns={"Clasificación General": "Clasificación_Comp"})
    df_resultado_ltvar = df_resultado_ltvar.merge(df_final[['SAP', 'Clasificación_Comp']], 
                                      left_on='Codigo_SAP', 
                                      right_on='SAP', 
                                      how='left')
    df_resultado_ltvar = df_resultado_ltvar.drop(columns=["SAP"])
    df_resultado_ltvar["Clasificación General"] = df_resultado_ltvar["Clasificación_Comp"].fillna(df_resultado_ltvar["Clasificación General"])
    df_resultado_ltvar = df_resultado_ltvar.drop(columns=["Clasificación_Comp"])
    
    return df_resultado_ltvar

def forecast_comp(df_forecast,dfMRP_filter):
    
    # Asegúrate que los nombres coincidan
    df_rel = dfMRP_filter.copy()
    #df_rel = dfMRP.copy()
    df_fc = df_forecast.copy()
    
    # Merge mantener todos los forecast
    df_resultado = df_fc.merge(
        df_rel,
        how="left",
        left_on="SAP_Origen",
        right_on="Producto Final"
    )
    
    # Rellenos cuando no hay relación
    df_resultado["Q"] = df_resultado["Q"].fillna(1)
    df_resultado["Componente"] = df_resultado["Componente"].fillna(df_resultado["SAP_Origen"])
    
    #Conversion
    df_resultado["Forecast_Temp"]=df_resultado["Forecast"]*df_resultado["Q"]
    
    #Agrupar
    df_agrupado = (
    df_resultado
    .groupby(["Componente", "Fecha"], as_index=False)["Forecast_Temp"]
    .sum()
    )
    
    # Cambiar nombres
    df_agrupado = df_agrupado.rename(columns={
    "Componente": "SAP_Origen",
    "Forecast_Temp": "Forecast"
    })
     
    # Orden 
    df_agrupado = df_agrupado.sort_values(["SAP_Origen", "Fecha"])

    return df_agrupado

def KPI_forecast(df_resultado_ltvar,df_KPI):
    
    df_resultado_ltvar = df_resultado_ltvar.merge(
        df_KPI[['Componente', '% Precision', '% MAPE' , 'RMSE' ,'BIAS']], 
        left_on='Codigo_SAP', 
        right_on='Componente', 
        how='left'
    )
    df_resultado_ltvar = df_resultado_ltvar.drop(columns=["Componente"])
    
    cols_metricas = ['% Precision', '% MAPE', 'RMSE', 'BIAS']

    mask = df_resultado_ltvar['Clasificación'].isin([
        'Estrategico',
        'Revision Periodica',
        'Gestion Simplificada'
    ])
    
    df_resultado_ltvar.loc[~mask, cols_metricas] = None
    
    return df_resultado_ltvar

def KPI_forecast_mespasado(df_resultado_ltvar,df_KPI):
    
    df_resultado_ltvar = df_resultado_ltvar.merge(
        df_KPI[['Componente', '% Precision', '% MAPE' , 'RMSE' ,'BIAS']], 
        left_on='Codigo_SAP', 
        right_on='Componente', 
        how='left'
    )
    df_resultado_ltvar = df_resultado_ltvar.drop(columns=["Componente"])
    
    cols_metricas = ['% Precision', '% MAPE', 'RMSE', 'BIAS']

    mask = df_resultado_ltvar['Clasificación'].isin([
        'Estrategico',
        'Revision Periodica',
        'Gestion Simplificada'
    ])
    
    df_resultado_ltvar.loc[~mask, cols_metricas] = None
    
    return df_resultado_ltvar

def calcular_Consumo_Mes_Actual_comp(dfActiveCode, dfBOMfinal):
    #dfActiveCode=df_resultado_ltvar.copy()
    # Limpieza y conversión de columnas numéricas
    dfActiveCode["Consumo Mes Actual"] = (
        dfActiveCode["Consumo Mes Actual"]
        .astype(str)
        .str.replace(',', '', regex=False)
        .replace(r'^\s*$', None, regex=True)
        .astype(float)
    )
    
    #Ajustamos compromentidos para componentes con setting   
    dfComprometido = dfActiveCode[['Codigo_SAP', 'Consumo Mes Actual']]

    comprometido_equivalente = dfComprometido.merge(
        dfBOMfinal, left_on="Codigo_SAP", right_on="Codigo_SAP", how="inner"
    )

    # Calcular el comprometido equivalente
    comprometido_equivalente["Consumo Mes Actual-C"] = (
        comprometido_equivalente["Consumo Mes Actual"] * comprometido_equivalente["Q"]
    )

    # Agrupar por Componente para acumular el tránsito total de cada componente
    comprometido_por_componente = (
        comprometido_equivalente.groupby("Componente")["Consumo Mes Actual-C"]
        .sum()
        .reset_index()
    )

    # Renombrar columna para hacer merge con dfActiveCode
    comprometido_por_componente.rename(columns={"Componente": "Codigo_SAP"}, inplace=True)

    dfActiveCode = dfActiveCode.merge(
        comprometido_por_componente, on="Codigo_SAP", how="left"
    )

    # Rellenar NaNs con 0
    dfActiveCode["Consumo Mes Actual-C"] = dfActiveCode["Consumo Mes Actual-C"].fillna(0)

    return dfActiveCode


def calcular_cobertura_meses_relativo(
    df_forecast,
    fecha_inicio,
    valor_unidades,
    cobertura_maxima=36
):

    # COPIA
    df = df_forecast.copy()

    # TIPOS
    df["Fecha"] = pd.to_datetime(
        df["Fecha"],
        errors="coerce"
    )

    df["Forecast"] = pd.to_numeric(
        df["Forecast"],
        errors="coerce"
    )

    fecha_inicio = pd.to_datetime(
        fecha_inicio,
        errors="coerce"
    )

    # VALIDAR FECHA
    if pd.isna(fecha_inicio):
        return np.nan

    # FECHA ACTUAL
    hoy = datetime.today()

    # DATA DEL SAP
    df_sap = df

    # SIN DATA
    if df_sap.empty:
        return np.nan

    # PROMEDIO GLOBAL DEL FORECAST
    forecast_promedio = (
        df_sap["Forecast"]
        .fillna(0)
        .mean()
    )

    # DATA DESDE FECHA INICIO
    df_calculo = df_sap[
        df_sap["Fecha"] >= fecha_inicio
    ].copy()

    # SIN DATA
    if df_calculo.empty:
        return np.nan

    # SI TODO EL FORECAST ES 0
    if (
        df_calculo["Forecast"]
        .fillna(0)
        .sum()
    ) == 0:

        return cobertura_maxima

    # VARIABLES
    stock = valor_unidades
    cobertura = 0

    # RECORRER FORECAST
    for row in df_calculo.itertuples():

        consumo = row.Forecast
        fecha_mes = row.Fecha

        # EVITAR NaN
        if pd.isna(consumo):
            continue

        # VALIDAR SI ES MES ACTUAL
        es_mes_actual = (
            fecha_mes.month == hoy.month
            and fecha_mes.year == hoy.year
        )

        # FRACCION DEL MES
        if es_mes_actual:

            dias_mes = monthrange(
                hoy.year,
                hoy.month
            )[1]

            fraccion_mes = (
                hoy.day / dias_mes
            )

        else:

            fraccion_mes = 1

        # FORECAST 0
        # SI CUENTA COMO COBERTURA
        if consumo == 0:

            cobertura += fraccion_mes
            continue

        # COBERTURA COMPLETA
        if stock >= consumo:

            stock -= consumo
            cobertura += fraccion_mes

        # COBERTURA PARCIAL
        else:

            cobertura += (
                (stock / consumo)
                * fraccion_mes
            )

            stock = 0
            break

    # STOCK SOBRANTE
    # USAR PROMEDIO GLOBAL
    if (
        stock > 0
        and forecast_promedio > 0
    ):

        cobertura += (
            stock / forecast_promedio
        )

    # CAP MAXIMO
    cobertura = min(
        cobertura,
        cobertura_maxima
    )

    return round(cobertura, 2)



def generar_reporte_simulacion(
    df_base_simulacion,
    df_ingresos_f,
    df_forecast
):

    # =========================================================
    # FILTROS INICIALES
    # =========================================================

    saps_validos = set(df_base_simulacion["Codigo_SAP"])

    df_ingresos_f = df_ingresos_f[
        df_ingresos_f["Número de artículo"].isin(saps_validos)
    ].copy()

    df_forecast = df_forecast[
        df_forecast["SAP_Origen"].isin(saps_validos)
    ].copy()

    # =========================================================
    # COPIAS
    # =========================================================

    df_base = df_base_simulacion.copy()
    df_in = df_ingresos_f.copy()
    df_fc = df_forecast.copy()

    # =========================================================
    # TIPOS
    # =========================================================

    df_fc["Fecha"] = pd.to_datetime(df_fc["Fecha"])
    df_in["Mes_año"] = pd.to_datetime(df_in["Mes_año"])

    # =========================================================
    # AGRUPACIONES
    # =========================================================

    forecast_grouped = {
        sap: grupo.sort_values("Fecha")
        for sap, grupo in df_fc.groupby("SAP_Origen")
    }

    ingresos_grouped = {
        sap: grupo
        for sap, grupo in df_in.groupby("Número de artículo")
    }

    # =========================================================
    # RESULTADO FINAL
    # =========================================================

    filas_finales = []

    # =========================================================
    # CONTROL FILAS EXCEL
    # =========================================================

    fila_excel_actual = 1

    # =========================================================
    # LOOP PRINCIPAL
    # =========================================================

    for row in df_base.itertuples(index=False):

        # =====================================================
        # DATOS BASE
        # =====================================================

        sap = getattr(row, "Codigo_SAP")
        clasificacion = getattr(row, "Clasificacion")
        costo = getattr(row, "Costo")
        stock_inicial = getattr(row, "Stock")
        consumo_actual = getattr(row, "Consumo_Mes_Actual")

        # =====================================================
        # FORECAST
        # =====================================================

        if sap in forecast_grouped:

            df_fc_sap = forecast_grouped[sap]

            fechas_horizonte = (
                df_fc_sap["Fecha"]
                .drop_duplicates()
                .sort_values()
                .tolist()
            )

            horizonte_set = set(fechas_horizonte)

            fechas_texto = [
                fecha.strftime("%d/%m/%Y")
                for fecha in fechas_horizonte
            ]

            dict_out = dict(
                zip(
                    df_fc_sap["Fecha"],
                    df_fc_sap["Forecast"]
                )
            )

        else:

            df_fc_sap = pd.DataFrame()

            fechas_horizonte = []
            horizonte_set = set()
            fechas_texto = []

            dict_out = {}

        # =====================================================
        # INGRESOS
        # =====================================================

        if sap in ingresos_grouped:

            df_in_sap = ingresos_grouped[sap]

            df_in_sap = df_in_sap[
                df_in_sap["Mes_año"].isin(horizonte_set)
            ]

            dict_in = (
                df_in_sap
                .groupby("Mes_año")["Cantidad - C"]
                .sum()
                .to_dict()
            )

        else:

            dict_in = {}

        # =====================================================
        # FILAS
        # =====================================================

        fila_si = []
        fila_in = []
        fila_out = []
        fila_sf = []
        fila_cob = []
        fila_acumulado_out = []

        # =====================================================
        # POSICIONES EXCEL
        # =====================================================

        fila_si_excel = fila_excel_actual + 7
        fila_in_excel = fila_excel_actual + 8
        fila_out_excel = fila_excel_actual + 9
        fila_sf_excel = fila_excel_actual + 10
        fila_cob_excel = fila_excel_actual + 11
        fila_acumulado_excel = fila_excel_actual + 12

        # =====================================================
        # LOOP MESES
        # =====================================================

        stock_actual_numerico = stock_inicial

        total_meses = len(fechas_horizonte)

        for i, fecha in enumerate(fechas_horizonte):

            ingreso = dict_in.get(fecha, 0)
            salida = dict_out.get(fecha, 0)

            # =================================================
            # COLUMNAS EXCEL
            # =================================================

            col_excel_num = i + 2
            col_excel = get_column_letter(col_excel_num)

            # =================================================
            # SI
            # =================================================

            if i == 0:

                si = stock_inicial

                si_numerico = stock_inicial

            else:

                col_anterior = get_column_letter(
                    col_excel_num - 1
                )

                si = f"={col_anterior}{fila_sf_excel}"

                si_numerico = stock_actual_numerico

            # =================================================
            # SF
            # =================================================

            sf_numerico = (
                si_numerico
                + ingreso
                - salida
            )

            sf = (
                f"={col_excel}{fila_si_excel}"
                f"+{col_excel}{fila_in_excel}"
                f"-{col_excel}{fila_out_excel}"
            )

            # =================================================
            # ACUMULADO OUT
            # =================================================

            fila_out_inicio = get_column_letter(2)
            acumulado_formula = (
                f"=SUMA(${fila_out_inicio}${fila_out_excel}:{col_excel}{fila_out_excel})"
            )

            # =================================================
            # COBERTURA
            # =================================================
            
            col_fin_dinamico = get_column_letter(
                total_meses + 60 + 1
            )
            
            # ================================================
            # COLUMNA INICIO
            # ================================================
            
            col_inicio = get_column_letter(
                col_excel_num + 1
            )
            
            # ================================================
            # COLUMNA PREVIA
            # ================================================
            
            col_prev = get_column_letter(
                col_excel_num
            )
            
            # ================================================
            # RANGOS
            # ================================================
            
            rango_acum = (
                f"{col_inicio}{fila_acumulado_excel}:"
                f"{col_fin_dinamico}{fila_acumulado_excel}"
            )
            
            rango_out = (
                f"{col_inicio}{fila_out_excel}:"
                f"{col_fin_dinamico}{fila_out_excel}"
            )
            
            # ================================================
            # CONDICION
            # ================================================
            
            condicion = (
                f"{rango_acum}-{col_prev}{fila_acumulado_excel}"
                f">={col_excel}{fila_sf_excel}"
            )
            
            coincidencia = (
                f"COINCIDIR(VERDADERO;{condicion};0)"
            )
            
            # ================================================
            # FORMULA COBERTURA
            # ================================================
            
            cob_formula = f"""=SI(
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            INDICE({rango_acum};1)-{col_prev}{fila_acumulado_excel}={col_excel}{fila_sf_excel};
            
            1;
            
            ({col_excel}{fila_sf_excel}/INDICE({rango_out};1))
            
            );
            
            1
            
            );
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            INDICE({rango_acum};1)-{col_prev}{fila_acumulado_excel}={col_excel}{fila_sf_excel};
            
            1;
            
            ({col_excel}{fila_sf_excel}/INDICE({rango_out};1))
            
            );
            
            1
            
            )
            
            +
            
            SI(
            
            {col_excel}{fila_sf_excel}=(
            INDICE({rango_acum};{coincidencia})
            -{col_prev}{fila_acumulado_excel}
            );
            
            1;
            
            (
            
            {col_excel}{fila_sf_excel}
            
            -
            
            (
            
            INDICE(
            {rango_acum};
            {coincidencia}-1
            )
            
            -{col_prev}{fila_acumulado_excel}
            
            )
            
            )
            
            /
            
            INDICE(
            {rango_out};
            {coincidencia}
            )
            
            )
            
            +
            
            SI(
            
            {coincidencia}-2<0;
            
            0;
            
            {coincidencia}-2
            
            )
            
            )
            
            <0;
            
            0;
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            INDICE({rango_acum};1)-{col_prev}{fila_acumulado_excel}={col_excel}{fila_sf_excel};
            
            1;
            
            ({col_excel}{fila_sf_excel}/INDICE({rango_out};1))
            
            );
            
            1
            
            );
            
            SI(
            
            {coincidencia}=1;
            
            SI(
            
            INDICE({rango_acum};1)-{col_prev}{fila_acumulado_excel}={col_excel}{fila_sf_excel};
            
            1;
            
            ({col_excel}{fila_sf_excel}/INDICE({rango_out};1))
            
            );
            
            1
            
            )
            
            +
            
            SI(
            
            {col_excel}{fila_sf_excel}=(
            INDICE({rango_acum};{coincidencia})
            -{col_prev}{fila_acumulado_excel}
            );
            
            1;
            
            (
            
            {col_excel}{fila_sf_excel}
            
            -
            
            (
            
            INDICE(
            {rango_acum};
            {coincidencia}-1
            )
            
            -{col_prev}{fila_acumulado_excel}
            
            )
            
            )
            
            /
            
            INDICE(
            {rango_out};
            {coincidencia}
            )
            
            )
            
            +
            
            SI(
            
            {coincidencia}-2<0;
            
            0;
            
            {coincidencia}-2
            
            )
            
            )
            
            )"""


            # =================================================
            # GUARDADO
            # =================================================

            fila_si.append(si)

            fila_in.append(ingreso)

            fila_out.append(salida)

            fila_sf.append(sf)

            fila_cob.append(cob_formula)

            fila_acumulado_out.append(acumulado_formula)

            # =================================================
            # UPDATE STOCK
            # =================================================

            stock_actual_numerico = sf_numerico

        # =====================================================
        # PROMEDIOS
        # =====================================================

        horizonte_promedios = 60

        col_inicio_fijo = get_column_letter(2)
        col_fin_fijo = get_column_letter(total_meses)
        
        formula_promedio_fijo = (
            f'=PROMEDIO({col_inicio_fijo}{fila_out_excel}:{col_fin_fijo}{fila_out_excel})'
        )
        
        for j in range(horizonte_promedios):
        
            fila_out.append(formula_promedio_fijo)

            col_prom = get_column_letter(
                total_meses + 2 + j
            )

            acumulado_formula = (
                f"=SUMA($B${fila_out_excel}:{col_prom}{fila_out_excel})"
            )

            fila_acumulado_out.append(acumulado_formula)

            fila_si.append("")
            fila_in.append("")
            fila_sf.append("")
            fila_cob.append("")

        # =====================================================
        # BLOQUE FINAL
        # =====================================================

        bloque = [

            ["Codigo_SAP", sap],
            ["Clasificación", clasificacion],
            ["Costo S/.", costo],
            ["Stock Disponible", stock_inicial],
            ["Consumo mes actual", consumo_actual],

            [],

            [""] + fechas_texto,

            ["SI"] + fila_si,
            ["IN"] + fila_in,
            ["OUT"] + fila_out,
            ["SF"] + fila_sf,
            ["COB"] + fila_cob,
            ["ACUM OUT"] + fila_acumulado_out,

            [],
            []

        ]

        filas_finales.extend(bloque)

        # =====================================================
        # UPDATE FILA EXCEL
        # =====================================================

        fila_excel_actual += len(bloque)

    # =========================================================
    # DF FINAL
    # =========================================================

    df_final = pd.DataFrame(filas_finales)

    return df_final

def cobertura_despues_LT(df_resultado_ltvar, df_forecast, columna_dinamica):

    #columna_dinamica="Sugerencia de Compra"
    hoy = datetime.today().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    df_forecast = df_forecast.copy()
    df_resultado_ltvar = df_resultado_ltvar.copy()

    df_forecast["Fecha"] = pd.to_datetime(df_forecast["Fecha"], errors="coerce")
    df_forecast["Forecast"] = pd.to_numeric(df_forecast["Forecast"], errors="coerce")

    df_forecast = df_forecast.sort_values(["SAP_Origen", "Fecha"])

    resultados = []

    for sap, grupo in df_forecast.groupby("SAP_Origen"):

        fila = df_resultado_ltvar.loc[df_resultado_ltvar["Codigo_SAP"] == sap]

        if fila.empty:
            resultados.append((sap, np.nan))
            continue

        stock = fila[columna_dinamica].iloc[0]
        lt_meses = fila["Promedio_LT"].iloc[0]

        if pd.isna(stock) or pd.isna(lt_meses):
            resultados.append((sap, np.nan))
            continue

        #  CONVERTIR LT EN MESES A FECHA
        meses_enteros = int(lt_meses)
        fraccion_mes = lt_meses - meses_enteros
        fecha_inicio = hoy.replace(day=1)

        # sumar meses enteros
        mes = fecha_inicio.month - 1 + meses_enteros
        año = fecha_inicio.year + mes // 12
        mes = mes % 12 + 1
        fecha_inicio = fecha_inicio.replace(year=año, month=mes)

        # ajuste por fracción del mes (en días)
        if fraccion_mes > 0:
            dias_mes = monthrange(fecha_inicio.year, fecha_inicio.month)[1]
            dias_extra = int(dias_mes * fraccion_mes)
            fecha_inicio = fecha_inicio.replace(day=1) + pd.Timedelta(days=dias_extra)

        primer_dia_mes = fecha_inicio.replace(day=1)

        grupo = grupo[grupo["Fecha"] >= primer_dia_mes].copy()

        if grupo.empty:
            resultados.append((sap, np.nan))
            continue

        dias_mes = monthrange(fecha_inicio.year, fecha_inicio.month)[1]

        fraccion_primer_mes = (
            dias_mes - fecha_inicio.day + 1
        ) / dias_mes

        cobertura = 0
        forecast_promedio = grupo["Forecast"].sum() / 12
        primer_mes = True

        for row in grupo.itertuples():

            consumo = row.Forecast

            if pd.isna(consumo):
                continue

            if primer_mes:

                if consumo == 0:
                    cobertura += fraccion_primer_mes

                elif stock >= consumo:
                    stock -= consumo
                    cobertura += fraccion_primer_mes

                else:
                    cobertura += (stock / consumo) * fraccion_primer_mes
                    stock = 0
                    break

                primer_mes = False

            else:

                if consumo == 0:
                    cobertura += 1
                    continue

                if stock >= consumo:
                    stock -= consumo
                    cobertura += 1

                else:
                    cobertura += stock / consumo
                    stock = 0
                    break

        if stock > 0 and forecast_promedio > 0:
            cobertura += stock / forecast_promedio

        resultados.append((sap, cobertura))

    df_cobertura = pd.DataFrame(
        resultados,
        columns=["Codigo_SAP", "Meses_Forecast"]
    )

    return df_resultado_ltvar.merge(
        df_cobertura,
        on="Codigo_SAP",
        how="left"
    )

def ajustar_multiplo_ensamble(df, columna, multiplo_ensamble):

    if Linea_Negocio != "EQUIPOS TRANS. MATER":
        return df

    for idx, row in df.iterrows():

        sap = row["Codigo_SAP"]

        if sap not in multiplo_ensamble:
            continue

        multiplo = pd.to_numeric(multiplo_ensamble[sap], errors="coerce")

        if pd.isna(multiplo) or multiplo <= 0:
            continue

        valor = pd.to_numeric(row[columna], errors="coerce")

        if pd.notna(valor):
            df.at[idx, columna] = int(np.ceil(valor / multiplo) * multiplo)

    return df

